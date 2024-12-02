import tkinter as tk
from tkinter import ttk, messagebox
import json
import os
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, FrameBreak, PageTemplate, BaseDocTemplate, Frame, Image as ReportLabImage
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
import tempfile
import subprocess
import platform
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from edit_report_page import EditReportPage
import shutil
from reportlab.lib.utils import ImageReader
from PIL import Image as PILImage
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
import uuid

thin_border = Side(border_style="thin", color="000000")  # 细线
medium_border = Side(border_style="medium", color="000000")  # 中等粗线
thick_border = Side(border_style="thick", color="000000")  # 粗线
double_border = Side(border_style="double", color="000000")  # 双线
dashed_border = Side(border_style="dashed", color="000000")  # 虚线

def is_dict_empty(d):
    """
    检查字典中的所有值是否为空
    空的定义: None, "", [], {}, 0, "0", "未知", "无"
    返回: True 如果所有值都为空，False 如果存在非空值
    """
    empty_values = [None, "", [], {}, 0, "0", "未知", "无"]
    return all(value in empty_values or (isinstance(value, str) and value.strip() == "") 
              for value in d.values())
    
def set_merged_cell_border(ws, range_string, border):
    if ':' in range_string:
        rows = ws[range_string]
        for row in rows:
            for cell in row:
                cell.border = border
    else:
        ws[range_string].border = border
            

def set_section_title(ws, cell_anchor, title):
    
    range_string = 'A' + cell_anchor + ":" + 'M' +cell_anchor
    start_cell = 'A' + cell_anchor
    
    ws.merge_cells(range_string)
    ws[start_cell] = title
    ws[start_cell].alignment = Alignment(horizontal='left')

def set_cell_element(ws, range_string, element, color=None, border=None):
    
    '''
    ws: excel
    range_strng: 'A1:C1' like
    element: string or value or something
    '''
    
    if ':' in range_string:
    
        ws.merge_cells(range_string)
        start_cell = range_string.split(':')[0]
        
    else:
        start_cell = range_string
    
    if color is not None:
        if ':' in range_string:
            cells = ws[range_string]
            for row in cells:
                for cell in row:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        else:
            ws[start_cell].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                
    if border is not None:
        set_merged_cell_border(ws, range_string, border)
        
    ws[start_cell] = element

class MyDocTemplate(BaseDocTemplate):
    def __init__(self, filename, **kw):
        BaseDocTemplate.__init__(self, filename, **kw)
        frame = Frame(
            self.leftMargin, self.bottomMargin, 
            self.width, self.height, 
            id='normal'
        )
        template = PageTemplate('normal', [frame], onPage=self.add_page_elements)
        self.addPageTemplates([template])

    def add_page_elements(self, canvas, doc):
        canvas.saveState()
        canvas.setFont(self.font_name, 10)
        canvas.drawRightString(doc.pagesize[0] - 1*cm, 1*cm, f"检查者: {self.examiner}")
        canvas.restoreState()

class CustomDocTemplate(MyDocTemplate):
    def __init__(self, filename, examiner, font_name, **kw):
        self.examiner = examiner
        self.font_name = font_name
        MyDocTemplate.__init__(self, filename, **kw)

class DatabasePage(ttk.Frame):
    def __init__(self, master, controller):
        super().__init__(master)
        self.controller = controller
        self.db_path = self.controller.db_path
        self.config_file = self.controller.config_file
        self.load_config()
        self.create_widgets()


    def create_widgets(self):
        # 创建主框架
        main_frame = ttk.Frame(self, padding="20 20 20 20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        # 创建搜索框
        search_frame = ttk.Frame(main_frame)
        search_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        ttk.Label(search_frame, text="搜索:").grid(row=0, column=0, padx=(0, 5))
        self.search_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.search_var, width=30).grid(row=0, column=1, padx=(0, 5))
        ttk.Button(search_frame, text="搜索", command=self.search_reports).grid(row=0, column=2)

        # 创建报告列表
        self.report_tree = ttk.Treeview(main_frame, columns=('ID', '姓名', '检查时间'), show='headings')
        self.report_tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.report_tree.heading('ID', text='患者ID')
        self.report_tree.heading('姓名', text='姓名')
        self.report_tree.heading('检查时间', text='检查时间')

        # 添加滚动条
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.report_tree.yview)
        scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.report_tree.configure(yscrollcommand=scrollbar.set)

        # 创建按钮架
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        ttk.Button(button_frame, text="查看详情", command=self.view_report).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(button_frame, text="编辑报告", command=self.edit_report).grid(row=0, column=1, padx=(0, 5))
        ttk.Button(button_frame, text="删除报告", command=self.delete_report).grid(row=0, column=2, padx=(0, 5))
        ttk.Button(button_frame, text="生成HIS文件", command=self.generate_his_files).grid(row=0, column=3, padx=(0, 5))
        ttk.Button(button_frame, text="刷新列表", command=self.load_reports).grid(row=0, column=4)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.report_tree.yview)
        scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.report_tree.configure(yscrollcommand=scrollbar.set)

        # 加载报告数据
        self.load_reports()

    def load_reports(self):
        self.report_tree.delete(*self.report_tree.get_children())  # 清空现有的报告列表
        report_folder = os.path.join(self.db_path, "report")
        
        for date_folder in os.listdir(report_folder):
            date_path = os.path.join(report_folder, date_folder)
            if os.path.isdir(date_path):
                for report_file in os.listdir(date_path):
                    if report_file.endswith('.json'):
                        file_path = os.path.join(date_path, report_file)
                        with open(file_path, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                        
                        basic_info = data.get("基本信息", {})
                        patient_id = basic_info.get("ID", "未知")
                        name = basic_info.get("姓名", "未知")
                        exam_time = basic_info.get("检查时间", "未知")
                        
                        self.report_tree.insert("", "end", values=(patient_id, name, exam_time), tags=(file_path,))

        # 按检查时间排序
        # self.report_tree.set_children('', sorted(self.report_tree.get_children(''), key=lambda x: self.report_tree.item(x)['values'][2], reverse=True))


    def generate_his_files(self):
        # 创建HIS文件夹
        his_dir = os.path.join(self.db_path, "HIS")
        # 删除HIS文件里的所有内容
        shutil.rmtree(his_dir, ignore_errors=True)
        os.makedirs(his_dir)
        
        # 获取数据库中的所有报告文件
        report_files = []
        for root, dirs, files in os.walk(os.path.join(self.db_path, "report")):
            for file in files:
                if file.endswith('.json'):
                    report_files.append(os.path.join(root, file))
        
        success_count = 0
        error_count = 0
        
        # 创建进度条窗口
        progress_window = tk.Toplevel(self)
        progress_window.title("生成HIS文件")
        progress_label = ttk.Label(progress_window, text="正在生成HIS文件...")
        progress_label.pack(pady=10)
        progress_bar = ttk.Progressbar(progress_window, length=300, mode='determinate')
        progress_bar.pack(pady=10)
        progress_bar['maximum'] = len(report_files)
        
        # 处理每个报告文件
        for i, file_path in enumerate(report_files):
            try:
                # 读取JSON文件
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                file_name = os.path.basename(file_path)
                file_name = file_name.replace('.json', '.txt')
                date_str = file_path.split('\\')[-2]
                
                # 生成HIS报告内容
                his_content = self.save_his_report(data)
                his_file_path = os.path.join(his_dir, date_str, file_name)
                
                # 写入文件
                os.makedirs(os.path.dirname(his_file_path), exist_ok=True)
                with open(his_file_path, 'w', encoding='utf-8') as f:
                    f.write(his_content)
                
                success_count += 1
                
            except Exception as e:
                print(f"处理文件 {file_path} 时出错: {str(e)}")
                error_count += 1
            
            # 更新进度条
            progress_bar['value'] = i + 1
            progress_window.update_idletasks()
        
        # 关闭进度条窗口
        progress_window.destroy()
        
        # 显示完成消息
        message = f"HIS文件生成完成！\n成功: {success_count} 个文件\n"
        if error_count > 0:
            message += f"失败: {error_count} 个文件"
        
        messagebox.showinfo("完成", message)
            
    
    def search_reports(self):
        # 实现搜索功能
        pass

    def calculate_age(self, birth_date):
        today = datetime.now()
        birth_date = datetime.strptime(birth_date, "%Y/%m/%d")
        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        return age

    def view_report(self):
        """生成检查报告"""
        selected_item = self.report_tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一个报告")
            return

        # 读取报告数据
        file_path = self.report_tree.item(selected_item)['tags'][0]
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        base_path = os.path.dirname(__file__)

        # 加载Excel模板
        template_path = os.path.join(base_path, "template", "report_template.xlsx")
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("错误", f"无法加载Excel模板: {e}")
            return

        ws = wb.worksheets[0]
        
        light_gray = "F0F0F0"
        
        border = Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border
        )

        # 填充基本信息
        basic_info = data.get("基本信息", {})
        ws['E3'] = basic_info.get("ID", "")
        ws['H3'] = basic_info.get("姓名", "")
        ws['K3'] = basic_info.get("性别", "")
        ws['E4'] = basic_info.get("出生日期", "")
        ws['H4'] = basic_info.get("检查时间", "")
        ws['K4'] = basic_info.get("检查设备", "")
        examine_doctor = basic_info.get("检查医生", "")
        
        cell_anchor = '6'

        # 填充自发性眼震动(spontaneous nystagmus)
        spontaneous_nystagmus = data.get("自发性眼震", {})
        if not is_dict_empty(spontaneous_nystagmus):
            
            set_section_title(ws, cell_anchor, '自发性眼震 (spontaneous nystagmus)')
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            element_cell = "A" + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, element_cell, '模式', color=light_gray)
            element_cell = 'D' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, element_cell, '速度（度/秒）', color=light_gray)
            element_cell = 'G' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, element_cell, '固视抑制', color=light_gray)
            element_cell = 'K' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, element_cell, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            element_cell = "A" + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, element_cell, spontaneous_nystagmus.get("自发性眼震模式", ""), border=border)
            element_cell = 'D' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, element_cell, spontaneous_nystagmus.get("自发性眼震速度", ""), border=border)
            element_cell = 'G' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, element_cell, spontaneous_nystagmus.get("自发性眼震固视抑制", ""), border=border)
            element_cell = 'K' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, element_cell, spontaneous_nystagmus.get("自发性眼震检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        # 摇头试验
        head_shake_test = data.get("摇头试验", {})
        if not is_dict_empty(head_shake_test):
            
            set_section_title(ws, cell_anchor, '摇头试验 (head-shaking test)')
    
            cell_anchor = str(int(cell_anchor) + 1)
            
            element_cell = "A" + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, element_cell, '模式', color=light_gray)
            element_cell = 'D' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, element_cell, '方向', color=light_gray)
            element_cell = 'G' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, element_cell, '速度（度/秒）', color=light_gray)
            element_cell = 'K' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, element_cell, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            element_cell = "A" + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, element_cell, head_shake_test.get("眼震模式", ""), border=border)
            element_cell = 'D' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, element_cell, head_shake_test.get("摇头方向", ""), border=border)
            element_cell = 'G' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, element_cell, head_shake_test.get("眼震速度", ""), border=border)
            element_cell = 'K' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, element_cell, head_shake_test.get("检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        
        # 凝视性眼震
        gaze_nystagmus = data.get("凝视性眼震", {})
        if not is_dict_empty(gaze_nystagmus):
            
            set_section_title(ws, cell_anchor, '凝视性眼震 (gaze-evoked nystagmus)')
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '凝视方向', color=light_gray)
            range_string = 'C' + cell_anchor + ':' + 'D' + cell_anchor
            set_cell_element(ws, range_string, '左', color=light_gray)
            range_string = 'E' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, '右', color=light_gray)
            range_string = 'G' + cell_anchor + ':' + 'H' + cell_anchor
            set_cell_element(ws, range_string, '上', color=light_gray)
            range_string = 'I' + cell_anchor + ':' + 'J' + cell_anchor
            set_cell_element(ws, range_string, '下', color=light_gray)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '眼震模式', color=light_gray)
            range_string = 'C' + cell_anchor + ':' + 'D' + cell_anchor
            set_cell_element(ws, range_string, gaze_nystagmus.get("凝视性眼震模式（左）", ""), border=border)
            range_string = 'E' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, gaze_nystagmus.get("凝视性眼震模式（右）", ""), border=border)
            range_string = 'G' + cell_anchor + ':' + 'H' + cell_anchor
            set_cell_element(ws, range_string, gaze_nystagmus.get("凝视性眼震模式（上）", ""), border=border)
            range_string = 'I' + cell_anchor + ':' + 'J' + cell_anchor
            set_cell_element(ws, range_string, gaze_nystagmus.get("凝视性眼震模式（下）", ""), border=border)
            range_string = 'L' + cell_anchor + ':' + 'M' + str(int(cell_anchor) + 1)
            set_cell_element(ws, range_string, gaze_nystagmus.get("凝视性眼震检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '眼震速度', color=light_gray)
            range_string = 'C' + cell_anchor + ':' + 'D' + cell_anchor
            set_cell_element(ws, range_string, gaze_nystagmus.get("凝视性眼震速度（左）", ""), border=border)
            range_string = 'E' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, gaze_nystagmus.get("凝视性眼震速度（右）", ""), border=border)
            range_string = 'G' + cell_anchor + ':' + 'H' + cell_anchor
            set_cell_element(ws, range_string, gaze_nystagmus.get("凝视性眼震速度（上）", ""), border=border)
            range_string = 'I' + cell_anchor + ':' + 'J' + cell_anchor
            set_cell_element(ws, range_string, gaze_nystagmus.get("凝视性眼震速度（下）", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        
        # 头脉冲试验
        head_impulse = data.get("头脉冲试验", "")
        if not is_dict_empty(head_impulse):
            
            
            set_section_title(ws, cell_anchor, '头脉冲试验 (head impulse test)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '侧别', color=light_gray)
            range_string = 'C' + cell_anchor
            set_cell_element(ws, range_string, '左外', color=light_gray)
            range_string = 'D' + cell_anchor
            set_cell_element(ws, range_string, '右外', color=light_gray)
            range_string = 'E' + cell_anchor
            set_cell_element(ws, range_string, '左前', color=light_gray)
            range_string = 'F' + cell_anchor
            set_cell_element(ws, range_string, '右后', color=light_gray)
            range_string = 'G' + cell_anchor
            set_cell_element(ws, range_string, '左后', color=light_gray)
            range_string = 'H' + cell_anchor
            set_cell_element(ws, range_string, '右前', color=light_gray)
            range_string = 'J' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, 'VOR增益', color=light_gray)
            range_string = 'C' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("VOR增益 (左外半规管)", ""), border=border)
            range_string = 'D' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("VOR增益 (右外半规管)", ""), border=border)
            range_string = 'E' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("VOR增益 (左前半规管)", ""), border=border)
            range_string = 'F' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("VOR增益 (右后半规管)", ""), border=border)
            range_string = 'G' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("VOR增益 (左后半规管)", ""), border=border)
            range_string = 'H' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("VOR增益 (右前半规管)", ""), border=border)
            range_string = 'J' + cell_anchor + ':' + 'M' + str(int(cell_anchor) + 2)
            head_impulse_result = head_impulse.get("头脉冲试验检查结果", [])
            head_impulse_result = ','.join(head_impulse_result)
            set_cell_element(ws, range_string, head_impulse_result, border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, 'PR分数', color=light_gray)
            range_string = 'C' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("PR分数 (左外半规管)", ""), border=border)
            range_string = 'D' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("PR分数 (右外半规管)", ""), border=border)
            range_string = 'E' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("PR分数 (左前半规管)", ""), border=border)
            range_string = 'F' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("PR分数 (右后半规管)", ""), border=border)
            range_string = 'G' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("PR分数 (左后半规管)", ""), border=border)
            range_string = 'H' + cell_anchor
            set_cell_element(ws, range_string, head_impulse.get("PR分数 (右前半规管)", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '扫视波', color=light_gray)
            
            sccade_wave = head_impulse.get("头脉冲试验扫视波", [])
            if '阴性' in sccade_wave or '配合欠佳' in sccade_wave:
                range_string = 'C' + cell_anchor + ':' + 'H' + cell_anchor
                start_cell = 'C' + cell_anchor
                ws.merge_cells(range_string)
                
                if '配合欠佳' in sccade_wave:
                    ws[start_cell] = '配合欠佳'
                else:
                    ws[start_cell] = '阴性'
                
            else:
                
                ws[f'C{cell_anchor}'].border = border
                ws[f'D{cell_anchor}'].border = border
                ws[f'E{cell_anchor}'].border = border
                ws[f'F{cell_anchor}'].border = border
                ws[f'G{cell_anchor}'].border = border
                ws[f'H{cell_anchor}'].border = border
                
                for i, option in enumerate(sccade_wave):
                    if option == "左外半规管":
                        ws[f'C{cell_anchor}'] = '√'
                    elif option == "右外半规管":
                        ws[f'D{cell_anchor}'] = '√'
                    elif option == "左前半规管":
                        ws[f'E{cell_anchor}'] = '√'
                    elif option == "右后半规管":
                        ws[f'F{cell_anchor}'] = '√'
                    elif option == "左后半规管":
                        ws[f'G{cell_anchor}'] = '√'
                    elif option == "右前半规管":
                        ws[f'H{cell_anchor}'] = '√'
                        
            cell_anchor = str(int(cell_anchor) + 2)
            
            # if head_impulse.get("头脉冲试验示意图") != "":
            #     pic_path = os.path.join(self.db_path, head_impulse.get("头脉冲试验示意图"))
            #     if os.path.exists(pic_path):
            #         img = openpyxl.drawing.image.Image(pic_path)
            #         img.anchor = 'O7'
                    
            #         cell_width = 6
            #         cell_height = 20
                    
            #         img.width = cell_width * 60
            #         img.height = cell_height * 18
                    
            #         ws.add_image(img)
            
        else:
            pass
        
        # 头脉冲抑制试验
        head_suppression = data.get("头脉冲抑制试验", "")
        if not is_dict_empty(head_suppression):
            
            set_section_title(ws, cell_anchor, '头脉冲抑制试验 (head impulse suppression)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '侧别', color=light_gray)
            range_string = 'C' + cell_anchor + ':' + 'D' + cell_anchor
            set_cell_element(ws, range_string, '增益', color=light_gray)
            range_string = 'E' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, '扫视波', color=light_gray)
            range_string = 'H' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '左外半规管', color=light_gray)
            range_string = 'C' + cell_anchor + ':' + 'D' + cell_anchor
            set_cell_element(ws, range_string, head_suppression.get("头脉冲抑制试验增益 (左外半规管)", ""), border=border)
            range_string = 'E' + cell_anchor + ':' + 'F' + cell_anchor
            ws.merge_cells(range_string)
            set_merged_cell_border(ws, range_string, border)
            
            
            range_string = 'H' + cell_anchor + ':' + 'M' + str(int(cell_anchor) + 1)
            set_cell_element(ws, range_string, head_suppression.get("头脉冲抑制试验检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '右外半规管', color=light_gray)
            range_string = 'C' + cell_anchor + ':' + 'D' + cell_anchor
            set_cell_element(ws, range_string, head_suppression.get("头脉冲抑制试验增益 (右外半规管)", ""), border=border)
            range_string = 'E' + cell_anchor + ':' + 'F' + cell_anchor
            ws.merge_cells(range_string)
            set_merged_cell_border(ws, range_string, border)
            
            
            
            sccade_wave = head_suppression.get("头脉冲抑制试验补偿性扫视波", [])
            if '阴性' in sccade_wave:
                ws[f'E{str(int(cell_anchor) - 1)}'] = '阴性'
                ws[f'E{cell_anchor}'] = '阴性'
            elif '配合欠佳' in sccade_wave:
                ws[f'E{str(int(cell_anchor) - 1)}'] = '配合欠佳'
                ws[f'E{cell_anchor}'] = '配合欠佳'
            
            else:
                if '左外半规管' in sccade_wave:
                    ws[f'E{str(int(cell_anchor) - 1)}'] = '√'
                if '右外半规管' in sccade_wave:
                    ws[f'E{cell_anchor}'] = '√'
                    
            cell_anchor = str(int(cell_anchor) + 2)
            
            # if head_suppression.get("头脉冲抑制试验示意图") != "":
            #     pic_path = os.path.join(self.db_path, head_suppression.get("头脉冲抑制试验示意图"))
            #     if os.path.exists(pic_path):
            #         img = openpyxl.drawing.image.Image(pic_path)
            #         img.anchor = 'O30'
                    
            #         cell_width = 6
            #         cell_height = 22
                    
            #         img.width = cell_width * 60
            #         img.height = cell_height * 18
                    
            #         ws.add_image(img)
        else:
            pass
        
        # 眼位反向偏斜（skew deviation）
        skew_deviation = data.get("眼位反向偏斜", "")
        if not is_dict_empty(skew_deviation):
            
            set_section_title(ws, cell_anchor, '眼位反向偏斜 (skew deviation)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, 'HR（度）', color=light_gray)
            range_string = 'D' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, 'VR（度）', color=light_gray)
            range_string = 'H' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, skew_deviation.get("眼位反向偏斜 (HR, 度)", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, skew_deviation.get("眼位反向偏斜 (VR, 度)", ""), border=border)
            range_string = 'H' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, skew_deviation.get("眼位反向偏斜检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        # 视觉增强前庭-眼反射试验 (VVOR)
        vvor = data.get("视觉增强前庭-眼反射试验", "")
        if not is_dict_empty(vvor):
            
            set_section_title(ws, cell_anchor, '视觉增强前庭-眼反射试验 (VVOR)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            range_string = 'C' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, vvor.get("检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        # 前庭-眼反射抑制试验（VOR suppression）
        vor_suppression = data.get("前庭-眼反射抑制试验", "")
        if not is_dict_empty(vor_suppression):
            
            set_section_title(ws, cell_anchor, '前庭-眼反射抑制试验 (VOR suppression)')
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            range_string = 'C' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, vor_suppression.get("检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        # 扫视检查
        pursuit_test = data.get("扫视检查", "")
        if not is_dict_empty(pursuit_test):
            
            set_section_title(ws, cell_anchor, '扫视检查 (pursuit test)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '视靶方向', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, '扫视延迟时间 (毫秒)', color=light_gray)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, '扫视峰速度 (度/秒)', color=light_gray)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, '扫视精确度 (%)', color=light_gray)
            range_string = 'I' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '左', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, pursuit_test.get("扫视延迟时间 (左向, 毫秒)", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, pursuit_test.get("扫视峰速度 (左向, 度/秒)", ""), border=border)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, pursuit_test.get("扫视精确度 (左向, %)", ""), border=border)
            
            range_string = 'I' + cell_anchor + ':' + 'M' + str(int(cell_anchor) + 1)
            set_cell_element(ws, range_string, pursuit_test.get("扫视检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '右', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, pursuit_test.get("扫视延迟时间 (右向, 毫秒)", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, pursuit_test.get("扫视峰速度 (右向, 度/秒)", ""), border=border)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, pursuit_test.get("扫视精确度 (右向, %)", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        # Dix-Hallpike试验
        dix_hallpike = data.get("位置试验 (Dix-Hallpike试验)", "")
        if not is_dict_empty(dix_hallpike):
            
            set_section_title(ws, cell_anchor, '位置试验 (Dix-Hallpike试验)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '侧别', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, '眼震模式', color=light_gray)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, '坐起眼震模式', color=light_gray)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, '出现眩晕/头晕', color=light_gray)
            range_string = 'H' + cell_anchor
            set_cell_element(ws, range_string, '潜伏期 (秒)', color=light_gray)
            range_string = 'I' + cell_anchor
            set_cell_element(ws, range_string, '持续时长 (秒)', color=light_gray)
            range_string = 'J' + cell_anchor + ':' + 'K' + cell_anchor
            set_cell_element(ws, range_string, '最大速度 (度/秒)', color=light_gray)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '眼震疲劳性', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '左', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("左侧眼震模式", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("左侧坐起眼震模式", ""), border=border)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("左侧出现眩晕/头晕", ""), border=border)
            range_string = 'H' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("左侧眼震潜伏期 (秒)", ""), border=border)
            range_string = 'I' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("左侧眼震持续时长 (秒)", ""), border=border)
            range_string = 'J' + cell_anchor + ':' + 'K' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("左侧眼震最大速度 (度/秒)", ""), border=border)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("左侧眼震疲劳性", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '右', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("右侧眼震模式", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("右侧坐起眼震模式", ""), border=border)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("右侧出现眩晕/头晕", ""), border=border)
            range_string = 'H' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("右侧眼震潜伏期 (秒)", ""), border=border)
            range_string = 'I' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("右侧眼震持续时长 (秒)", ""), border=border)
            range_string = 'J' + cell_anchor + ':' + 'K' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("右侧眼震最大速度 (度/秒)", ""), border=border)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("右侧眼震疲劳性", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            range_string = 'C' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, dix_hallpike.get("检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        # 仰卧滚转试验
        supine_roll = data.get("位置试验 (仰卧滚转试验)", "")
        if not is_dict_empty(supine_roll):
            
            set_section_title(ws, cell_anchor, '位置试验 (仰卧滚转试验)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '侧别', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, '眼震模式', color=light_gray)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, '出现眩晕/头晕', color=light_gray)
            range_string = 'F' + cell_anchor
            set_cell_element(ws, range_string, '潜伏期 (秒)', color=light_gray)
            range_string = 'G' + cell_anchor
            set_cell_element(ws, range_string, '持续时长 (秒)', color=light_gray)
            range_string = 'H' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, range_string, '最大速度 (度/秒)', color=light_gray)
            range_string = 'K' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '左', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, supine_roll.get("左侧眼震模式", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, supine_roll.get("左侧出现眩晕/头晕", ""), border=border)
            range_string = 'F' + cell_anchor
            set_cell_element(ws, range_string, supine_roll.get("左侧眼震潜伏期 (秒)", ""), border=border)
            range_string = 'G' + cell_anchor
            set_cell_element(ws, range_string, supine_roll.get("左侧眼震持续时长 (秒)", ""), border=border)
            range_string = 'H' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, range_string, supine_roll.get("左侧眼震最大速度 (度/秒)", ""), border=border)
            
            range_string = 'K' + cell_anchor + ':' + 'M' + str(int(cell_anchor) + 1)
            set_cell_element(ws, range_string, supine_roll.get("检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '右', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, supine_roll.get("右侧眼震模式", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, supine_roll.get("右侧出现眩晕/头晕", ""), border=border)
            range_string = 'F' + cell_anchor
            set_cell_element(ws, range_string, supine_roll.get("右侧眼震潜伏期 (秒)", ""), border=border)
            range_string = 'G' + cell_anchor
            set_cell_element(ws, range_string, supine_roll.get("右侧眼震持续时长 (秒)", ""), border=border)
            range_string = 'H' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, range_string, supine_roll.get("右侧眼震最大速度 (度/秒)", ""), border=border)

            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        
        # 位置试验（其他）
        other_position_test = data.get("位置试验(其他)", "")
        if not is_dict_empty(other_position_test):
            
            set_section_title(ws, cell_anchor, '位置试验 (其他)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '坐位-平卧试验', color=light_gray)
            range_string = 'C' + cell_anchor + ':' + 'D' + cell_anchor
            set_cell_element(ws, range_string, '坐位-低头试验', color=light_gray)
            range_string = 'E' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, '坐位-仰头试验', color=light_gray)
            range_string = 'G' + cell_anchor + ':' + 'H' + cell_anchor
            set_cell_element(ws, range_string, '零平面', color=light_gray)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, other_position_test.get("坐位-平卧试验", ""), border=border)
            range_string = 'C' + cell_anchor + ':' + 'D' + cell_anchor
            set_cell_element(ws, range_string, other_position_test.get("坐位-低头试验", ""), border=border)
            range_string = 'E' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, other_position_test.get("坐位-仰头试验", ""), border=border)
            range_string = 'G' + cell_anchor + ':' + 'H' + cell_anchor
            set_cell_element(ws, range_string, other_position_test.get("零平面", ""), border=border)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, other_position_test.get("检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        # 视跟踪
        visual_tracking = data.get("视跟踪", "")
        if not is_dict_empty(visual_tracking):
            
            set_section_title(ws, cell_anchor, '视跟踪 (visual tracking)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, '视跟踪曲线分型', color=light_gray)
            range_string = 'D' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, '视跟踪增益', color=light_gray)
            range_string = 'H' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, visual_tracking.get("视跟踪曲线分型", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'F' + cell_anchor
            set_cell_element(ws, range_string, visual_tracking.get("视跟踪增益", ""), border=border)
            range_string = 'H' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, visual_tracking.get("视跟踪检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)

            
        else:
            pass
        
        # 视动性眼震
        spontaneous_nystagmus = data.get("视动性眼震", "")
        if not is_dict_empty(spontaneous_nystagmus):
            
            set_section_title(ws, cell_anchor, '视动性眼震 (optokinetic Nystagmus)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '视靶方向', color=light_gray)
            range_string = 'B' + cell_anchor
            set_cell_element(ws, range_string, '左', color=light_gray)
            range_string = 'C' + cell_anchor
            set_cell_element(ws, range_string, '右', color=light_gray)
            range_string = 'D' + cell_anchor
            set_cell_element(ws, range_string, '上', color=light_gray)
            range_string = 'E' + cell_anchor
            set_cell_element(ws, range_string, '下', color=light_gray)
            range_string = 'G' + cell_anchor + ':' + 'H' + cell_anchor
            set_cell_element(ws, range_string, '视靶方向', color=light_gray)
            range_string = 'I' + cell_anchor
            set_cell_element(ws, range_string, '水平', color=light_gray)
            range_string = 'J' + cell_anchor
            set_cell_element(ws, range_string, '垂直', color=light_gray)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '增益', color=light_gray)
            range_string = 'B' + cell_anchor
            set_cell_element(ws, range_string, spontaneous_nystagmus.get("向左视标增益", ""), border=border)
            range_string = 'C' + cell_anchor
            set_cell_element(ws, range_string, spontaneous_nystagmus.get("向右视标增益", ""), border=border)
            range_string = 'D' + cell_anchor
            set_cell_element(ws, range_string, spontaneous_nystagmus.get("向上视标增益", ""), border=border)
            range_string = 'E' + cell_anchor
            set_cell_element(ws, range_string, spontaneous_nystagmus.get("向下视标增益", ""), border=border)
            
            range_string = 'G' + cell_anchor + ':' + 'H' + cell_anchor
            set_cell_element(ws, range_string, '不对称性 (%)', color=light_gray)
            range_string = 'I' + cell_anchor
            set_cell_element(ws, range_string, spontaneous_nystagmus.get("水平视标不对称性（%）", ""), border=border)
            range_string = 'J' + cell_anchor
            set_cell_element(ws, range_string, spontaneous_nystagmus.get("垂直视标不对称性（%）", ""), border=border)
            
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, spontaneous_nystagmus.get("检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        # 瘘管试验
        laceration_test = data.get("瘘管试验", "")
        if not is_dict_empty(laceration_test):
            
            set_section_title(ws, cell_anchor, '瘘管试验 (caloric test)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '侧别', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, '反应', color=light_gray)
            range_string = 'F' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '左', color=light_gray)
            range_string = 'A' + str(int(cell_anchor) + 1)
            set_cell_element(ws, range_string, '右', color=light_gray)
            
            positive_options = laceration_test.get("瘘管试验", [])
            
            
            if '配合欠佳' in positive_options:
                
                range_string = 'B' + cell_anchor + ':' + 'C' + str(int(cell_anchor) + 1)
                set_cell_element(ws, range_string, '配合欠佳', border=border)
                
            else:
                
                range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
                ws.merge_cells(range_string)
                set_merged_cell_border(ws, range_string, border)
                
                range_string = 'B' + str(int(cell_anchor) + 1) + ':' + 'C' + str(int(cell_anchor) + 1)
                ws.merge_cells(range_string)
                set_merged_cell_border(ws, range_string, border)
                
                if '阴性' in positive_options:
                    
                    ws[f'B{cell_anchor}'] = '阴性'
                    ws[f'B{str(int(cell_anchor) + 1)}'] = '阴性'
                        
                elif '双耳阳性' in positive_options:
                    ws[f'B{cell_anchor}'] = "阳性"
                    ws[f'B{str(int(cell_anchor) + 1)}'] = "阳性"
                    
                elif '双耳弱阳性' in positive_options:
                    ws[f'B{cell_anchor}'] = "弱阳性"
                    ws[f'B{str(int(cell_anchor) + 1)}'] = "弱阳性"
                
                elif '右耳阳性' in positive_options:
                    ws[f'B{str(int(cell_anchor) + 1)}'] = "阳性"
                    ws[f'B{cell_anchor}'] = "阴性"
                    
                elif '左耳阳性' in positive_options:
                    ws[f'B{str(int(cell_anchor) + 1)}'] = "阴性"
                    ws[f'B{cell_anchor}'] = "阳性"
                    
                elif '右耳弱阳性' in positive_options:
                    ws[f'B{str(int(cell_anchor) + 1)}'] = "弱阳性"
                    ws[f'B{cell_anchor}'] = "阴性"
                    
                elif '左耳弱阳性' in positive_options:
                    ws[f'B{str(int(cell_anchor) + 1)}'] = "阴性"
                    ws[f'B{cell_anchor}'] = "弱阳性"
                    
            range_string = 'F' + cell_anchor + ':' + 'M' + str(int(cell_anchor) + 1)
            set_cell_element(ws, range_string, laceration_test.get("检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 3)
            
        else:
            pass
        
        # 温度试验
        temperature_test = data.get("温度试验", "")
        if not is_dict_empty(temperature_test):
            
            set_section_title(ws, cell_anchor, '温度试验 (temperature test)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '单侧减弱侧别 (UW)', color=light_gray)
            range_string = 'C' + cell_anchor 
            set_cell_element(ws, range_string, temperature_test.get("单侧减弱侧别 (UW)", ""), border=border)
            
            range_string = 'E' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, '优势偏向侧别 (DP)', color=light_gray)
            range_string = 'H' + cell_anchor
            set_cell_element(ws, range_string, temperature_test.get("优势偏向侧别 (DP)", ""), border=border)
            
            range_string = 'J' + cell_anchor + ':' + 'L' + cell_anchor
            set_cell_element(ws, range_string, '最大慢相速度总和（右耳, 度/秒）', color=light_gray)
            range_string = 'M' + cell_anchor
            set_cell_element(ws, range_string, temperature_test.get("最大慢相速度总和（右耳, 度/秒）", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '单侧减弱数值 (UW, %)', color=light_gray)
            range_string = 'C' + cell_anchor
            set_cell_element(ws, range_string, temperature_test.get("单侧减弱数值 (UW, %)", ""), border=border)
            
            range_string = 'E' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, '优势偏向数值 (DP, 度/秒)', color=light_gray)
            range_string = 'H' + cell_anchor
            set_cell_element(ws, range_string, temperature_test.get("优势偏向数值 (DP, 度/秒)", ""), border=border)
            
            range_string = 'J' + cell_anchor + ':' + 'L' + cell_anchor
            set_cell_element(ws, range_string, '最大慢相速度总和（左耳, 度/秒）', color=light_gray)
            range_string = 'M' + cell_anchor
            set_cell_element(ws, range_string, temperature_test.get("最大慢相速度总和（左耳, 度/秒）", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, '固视抑制指数 (FI, %)', color=light_gray)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, temperature_test.get("固视抑制指数 (FI, %)", ""), border=border)
            
            range_string = 'G' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            range_string = 'J' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, temperature_test.get("检查结果", ""), border=border)
            
            
            cell_anchor = str(int(cell_anchor) + 2)
            
            # pic_path = temperature_test.get("温度试验示意图", "")
            # if pic_path != '':
            #     pic_path = os.path.join(self.db_path, head_impulse.get("头脉冲试验示意图"))
            #     if os.path.exists(pic_path):
            #         img = openpyxl.drawing.image.Image(pic_path)
            #         img.anchor = 'O58'
                    
            #         cell_width = 6
            #         cell_height = 16
                    
            #         img.width = cell_width * 60
            #         img.height = cell_height * 18
                    
            #         ws.add_image(img)
            
            
            
        else:
            pass
        
        # 颈肌前庭诱发肌源性电位
        cervical_evoked_myogenic_potential = data.get("颈肌前庭诱发肌源性电位 (cVEMP)", "")
        if not is_dict_empty(cervical_evoked_myogenic_potential):
            
            set_section_title(ws, cell_anchor, '颈肌前庭诱发肌源性电位 (cVEMP)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '侧别', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, '声强阈值 (分贝)', color=light_gray)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, 'P13波潜伏期 (毫秒)', color=light_gray)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, 'N23波潜伏期 (毫秒)', color=light_gray)
            range_string = 'H' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, range_string, 'P13-N23波间期 (毫秒)', color=light_gray)
            range_string = 'J' + cell_anchor + ':' + 'K' + cell_anchor
            set_cell_element(ws, range_string, 'P13波振幅 (微伏)', color=light_gray)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, 'N23波振幅 (微伏)', color=light_gray)
            range_string = 'N' + cell_anchor + ':' + 'O' + cell_anchor
            set_cell_element(ws, range_string, 'P13-N23波振幅 (微伏)', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '左', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("左耳声强阈值 (分贝)", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("左耳P13波潜伏期 (毫秒)", ""), border=border)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("左耳N23波潜伏期 (毫秒)", ""), border=border)
            range_string = 'H' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("左耳P13-N23波间期 (毫秒)", ""), border=border)
            range_string = 'J' + cell_anchor + ':' + 'K' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("左耳P13波振幅 (微伏)", ""), border=border)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("左耳N23波振幅 (微伏)", ""), border=border)
            range_string = 'N' + cell_anchor + ':' + 'O' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("左耳P13-N23波振幅 (微伏)", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '右', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("右耳声强阈值 (分贝)", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("右耳P13波潜伏期 (毫秒)", ""), border=border)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("右耳N23波潜伏期 (毫秒)", ""), border=border)
            range_string = 'H' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("右耳P13-N23波间期 (毫秒)", ""), border=border)
            range_string = 'J' + cell_anchor + ':' + 'K' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("右耳P13波振幅 (微伏)", ""), border=border)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("右耳N23波振幅 (微伏)", ""), border=border)
            range_string = 'N' + cell_anchor + ':' + 'O' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("右耳P13-N23波振幅 (微伏)", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, 'cVEMP耳间不对称性 (%)', color=light_gray)
            range_string = 'E' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("cVEMP耳间不对称性 (%)", ""), border=border)
            range_string = 'E' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, cervical_evoked_myogenic_potential.get("检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        # 眼肌前庭诱发肌源性电位 (oVEMP)
        ocular_evoked_myogenic_potential = data.get("眼肌前庭诱发肌源性电位 (oVEMP)", "")
        if not is_dict_empty(ocular_evoked_myogenic_potential):
            
            set_section_title(ws, cell_anchor, '眼肌前庭诱发肌源性电位 (oVEMP)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '侧别', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, '声强阈值 (分贝)', color=light_gray)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, 'N10波潜伏期 (毫秒)', color=light_gray)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, 'P15波潜伏期 (毫秒)', color=light_gray)
            range_string = 'H' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, range_string, 'N10-P15波间期 (毫秒)', color=light_gray)
            range_string = 'J' + cell_anchor + ':' + 'K' + cell_anchor
            set_cell_element(ws, range_string, 'N10波振幅 (微伏)', color=light_gray)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, 'P15波振幅 (微伏)', color=light_gray)
            range_string = 'N' + cell_anchor + ':' + 'O' + cell_anchor
            set_cell_element(ws, range_string, 'N10-P15波振幅 (微伏)', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '左', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("左耳声强阈值 (分贝)", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("左耳N10波潜伏期 (毫秒)", ""), border=border)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("左耳P15波潜伏期 (毫秒)", ""), border=border)
            range_string = 'H' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("左耳N10-P15波间期 (毫秒)", ""), border=border)
            range_string = 'J' + cell_anchor + ':' + 'K' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("左耳N10波振幅 (微伏)", ""), border=border)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("左耳P15波振幅 (微伏)", ""), border=border)
            range_string = 'N' + cell_anchor + ':' + 'O' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("左耳N10-P15波振幅 (微伏)", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor
            set_cell_element(ws, range_string, '右', color=light_gray)
            range_string = 'B' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("右耳声强阈值 (分贝)", ""), border=border)
            range_string = 'D' + cell_anchor + ':' + 'E' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("右耳N10波潜伏期 (毫秒)", ""), border=border)
            range_string = 'F' + cell_anchor + ':' + 'G' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("右耳P15波潜伏期 (毫秒)", ""), border=border)
            range_string = 'H' + cell_anchor + ':' + 'I' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("右耳N10-P15波间期 (毫秒)", ""), border=border)
            range_string = 'J' + cell_anchor + ':' + 'K' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("右耳N10波振幅 (微伏)", ""), border=border)
            range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("右耳P15波振幅 (微伏)", ""), border=border)
            range_string = 'N' + cell_anchor + ':' + 'O' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("右耳N10-P15波振幅 (微伏)", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, 'oVEMP耳间不对称性 (%)', color=light_gray)
            range_string = 'E' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'C' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("oVEMP耳间不对称性 (%)", ""), border=border)
            range_string = 'E' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, ocular_evoked_myogenic_potential.get("检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        # 主观视觉垂直线
        subjective_visual_vertical_line = data.get("主观视觉垂直线 (SVV)", "")
        if not is_dict_empty(subjective_visual_vertical_line):
            
            set_section_title(ws, cell_anchor, '主观视觉垂直线 (SVV)')
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, '偏斜方向', color=light_gray)
            range_string = 'C' + cell_anchor + ':' + 'D' + cell_anchor
            set_cell_element(ws, range_string, '偏斜角度（度）', color=light_gray)
            range_string = 'F' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, '检查结果', color=light_gray)
            
            cell_anchor = str(int(cell_anchor) + 1)
            
            range_string = 'A' + cell_anchor + ':' + 'B' + cell_anchor
            set_cell_element(ws, range_string, subjective_visual_vertical_line.get("偏斜方向", ""), border=border)
            range_string = 'C' + cell_anchor + ':' + 'D' + cell_anchor
            set_cell_element(ws, range_string, subjective_visual_vertical_line.get("偏斜角度（度）", ""), border=border)
            range_string = 'F' + cell_anchor + ':' + 'M' + cell_anchor
            set_cell_element(ws, range_string, subjective_visual_vertical_line.get("检查结果", ""), border=border)
            
            cell_anchor = str(int(cell_anchor) + 2)
            
        else:
            pass
        
        exp_seen = self.save_his_report_conclusion(data)
        range_string = 'A' + cell_anchor + ':' + 'C' + str(int(cell_anchor) + 2)
        set_cell_element(ws, range_string, '检查所见', color=light_gray)
        range_string = 'D' + cell_anchor + ':' + 'M' + str(int(cell_anchor) + 2)
        set_cell_element(ws, range_string, exp_seen, border=border)
        # set exp_seen cell left align and wrap text and shrink to fit
        first_cell = 'D' + cell_anchor
        ws[first_cell].alignment = Alignment(horizontal='left',  # 左对齐
                                           vertical='center',    # 垂直居中
                                           wrap_text=True)   
        
        ws[first_cell].font = Font(size=8)
        
        cell_anchor = str(int(cell_anchor) + 4)
        
        exp_result = data.get("检查结论", "")
        exp_result = ','.join(exp_result)
        range_string = 'A' + cell_anchor + ':' + 'C' + cell_anchor
        set_cell_element(ws, range_string, '检查印象', color=light_gray)
        range_string = 'D' + cell_anchor + ':' + 'M' + cell_anchor
        set_cell_element(ws, range_string, exp_result, border=border)
        
        cell_anchor = str(int(cell_anchor) + 2)
        
        range_string = 'J' + cell_anchor + ':' + 'K' + cell_anchor
        set_cell_element(ws, range_string, '检查医师', color=light_gray)
        range_string = 'L' + cell_anchor + ':' + 'M' + cell_anchor
        set_cell_element(ws, range_string, examine_doctor, border=border)
        
        # # 设置页面布局
        # ws.page_setup.paperSize = 9  # A4纸
        # ws.page_setup.orientation = 'landscape'  # 横向
        # ws.page_setup.fitToWidth = 1  # 调整为1页宽
        # ws.page_setup.fitToHeight = 1  # 调整为1页高
        
        # # 设置打印区域边距（单位：英寸）
        # ws.page_margins.left = 0.5
        # ws.page_margins.right = 0.5
        # ws.page_margins.top = 0.5
        # ws.page_margins.bottom = 0.5
        
        # 切换到第二页放图片
        
        ws = wb.worksheets[1]
        cell_anchor = '1'
        
        # 头脉冲试验示意图
        head_impulse = data.get("头脉冲试验", "")
        if not is_dict_empty(head_impulse):

            
            pic_path = head_impulse.get("头脉冲试验示意图", "")
            if pic_path != '': 
                
                pic_path = os.path.join(self.db_path, head_impulse.get("头脉冲试验示意图"))
                if platform.system() == "Darwin":
                    pic_path = pic_path.replace('\\', '/')
                if os.path.exists(pic_path):
                    
                    set_section_title(ws, cell_anchor, '头脉冲试验 (head impulse test)')
                    cell_anchor = str(int(cell_anchor) + 1)
                    
                    img = openpyxl.drawing.image.Image(pic_path)
                    img.anchor = 'A' + cell_anchor
                    
                    cell_width = 6
                    cell_height = 16
                    
                    img.width = cell_width * 60
                    img.height = cell_height * 18
                    
                    ws.add_image(img)
            
            cell_anchor = str(int(cell_anchor) + 20)
            
        else:
            pass
        
        # 头脉冲抑制试验示意图
        head_impulse_suppression = data.get("头脉冲抑制试验", "")
        if not is_dict_empty(head_impulse_suppression):
                
                pic_path = head_impulse_suppression.get("头脉冲抑制试验示意图", "")
                if pic_path != '':
                    pic_path = os.path.join(self.db_path, head_impulse_suppression.get("头脉冲抑制试验示意图"))
                    if platform.system() == "Darwin":
                        pic_path = pic_path.replace('\\', '/')
                    if os.path.exists(pic_path):
                        
                        set_section_title(ws, cell_anchor, '头脉冲抑制试验 (head impulse suppression test)')
                        cell_anchor = str(int(cell_anchor) + 1)
                        
                        img = openpyxl.drawing.image.Image(pic_path)
                        img.anchor = 'A' + cell_anchor
                        
                        cell_width = 6
                        cell_height = 16
                        
                        img.width = cell_width * 60
                        img.height = cell_height * 18
                        
                        ws.add_image(img)
                
                cell_anchor = str(int(cell_anchor) + 20)
                
        else:
            pass
        
        # 温度试验示意图
        temperature_test = data.get("温度试验", "")
        if not is_dict_empty(temperature_test):
            
            pic_path = temperature_test.get("温度试验示意图", "")
            if platform.system() == "Darwin":
                pic_path = pic_path.replace('\\', '/')
            if pic_path != '':
                pic_path = os.path.join(self.db_path, temperature_test.get("温度试验示意图"))
                if os.path.exists(pic_path):
                    set_section_title(ws, cell_anchor, '温度试验 (temperature test)')
                    cell_anchor = str(int(cell_anchor) + 1)
                    
                    img = openpyxl.drawing.image.Image(pic_path)
                    img.anchor = 'A1'
                    
                    cell_width = 6
                    cell_height = 16
                    
                    img.width = cell_width * 60
                    img.height = cell_height * 18
                    
                    ws.add_image(img)
            
            cell_anchor = str(int(cell_anchor) + 20)
            
        else:
            pass
        
        # 生成唯一的临时文件名
        random_id = str(uuid.uuid4())
        excel_path = os.path.join(tempfile.gettempdir(), f"report_{basic_info.get('ID', 'temp')}_{random_id}.xlsx")
        
        # 保存Excel文件
        wb.save(excel_path)
        
        # 直接打开Excel文件
        try:
            if platform.system() == "Windows":
                os.startfile(excel_path)
            elif platform.system() == "Darwin":  # macOS
                subprocess.call(["open", excel_path])
            else:  # Linux
                subprocess.call(["xdg-open", excel_path])
        except Exception as e:
            messagebox.showerror("错误", f"无法打开Excel文件: {str(e)}")

    def edit_report(self):
        selected_item = self.report_tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一个报告")
            return

        file_path = self.report_tree.item(selected_item)['tags'][0]
        edit_page = EditReportPage(self, file_path)
        self.wait_window(edit_page)
        # 刷新报告列表
        self.load_reports()

    def delete_report(self):
        selected_item = self.report_tree.selection()
        if not selected_item:
            messagebox.showwarning("警告", "请先选择一个报告")
            return

        if messagebox.askyesno("确认", "确定要删除选中的报告吗？"):
            file_path = self.report_tree.item(selected_item)['tags'][0]
            
            # 创建删除的归档目录
            deleted_report_dir = os.path.join(self.db_path, "arch", "deleted", "report")
            deleted_pic_dir = os.path.join(self.db_path, "arch", "deleted", "pic")
            os.makedirs(deleted_report_dir, exist_ok=True)
            os.makedirs(deleted_pic_dir, exist_ok=True)

            # 获取当前日期和时间
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

            # 移动报告文件
            report_filename = os.path.basename(file_path)
            new_report_filename = f"{os.path.splitext(report_filename)[0]}_{current_time}{os.path.splitext(report_filename)[1]}"
            shutil.move(file_path, os.path.join(deleted_report_dir, new_report_filename))

            # 移动相关图片
            with open(os.path.join(deleted_report_dir, new_report_filename), 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            for test_name in ['头脉冲试验', '头脉冲抑制试验']:
                if test_name in data and f'{test_name}示意图' in data[test_name]:
                    pic_path = data[test_name][f'{test_name}示意图']
                    if os.path.exists(os.path.join(self.db_path, pic_path)):
                        pic_filename = os.path.basename(pic_path)
                        new_pic_filename = f"{os.path.splitext(pic_filename)[0]}_{current_time}{os.path.splitext(pic_filename)[1]}"
                        shutil.move(os.path.join(self.db_path, pic_path), os.path.join(deleted_pic_dir, new_pic_filename))

            # 从树形视图中删除项目
            self.report_tree.delete(selected_item)

            messagebox.showinfo("成功", "报告已成功删除并归档")
    
    def save_his_report_conclusion(self, data):
        # 收集所有检查结果
        results = []
        
        # 定义检查项目及其对应的结果字段名
        test_items = {
            "自发性眼震": ["自发性眼震检查结果"],
            "凝视性眼震": ["凝视性眼震检查结果"],
            "头脉冲试验": ["头脉冲试验检查结果"],
            "头脉冲抑制试验": ["头脉冲抑制试验检查结果"],
            "眼位反向偏斜": ["眼位反向偏斜检查结果"],
            "扫视检查": ["扫视检查结果"],
            "视觉增强前庭-眼反射试验": ["检查结果"],
            "前庭-眼反射抑制试验": ["检查结果"],
            "摇头试验": ["检查结果"],
            "位置试验 (Dix-Hallpike试验)": ["检查结果"],
            "位置试验 (仰卧滚转试验)": ["检查结果"],
            "位置试验（其他）": ["检查结果"],
            "视跟踪": ["视跟踪检查结果"],
            "视动性眼震": ["检查结果"],
            "瘘管试验": ["瘘管试验"],
            "温度试验": ["检查结果"],
            "颈肌前庭诱发肌源性电位": ["检查结果"],
            "眼肌前庭诱发肌源性电位 (oVEMP)": ["检查结果"],
            "主观视觉垂直线 (SVV)": ["检查结果"]
        }
        
        # 遍历每个检查项目
        for item, result_fields in test_items.items():
            test_data = data.get(item, {})
            if not is_dict_empty(test_data):
                for field in result_fields:
                    result = test_data.get(field, "")
                    if result:
                        # 处理列表类型的结果
                        if isinstance(result, list):
                            result = "、".join(result)
                        results.append(f"{item}：{result}")
        
        # 结果中间用分号隔开
        return ';'.join(results)
    
    def save_his_report(self, data):
        
        # 收集所有检查结果
        results = []
        
        # 自发性眼震
        spontaneous = data.get("自发性眼震", {})
        if not is_dict_empty(spontaneous):
            if spontaneous.get("自发性眼震模式", "") != "":
                results.append("自发性眼震模式：" + spontaneous.get("自发性眼震模式", ""))
            if spontaneous.get("自发性眼震速度", "") != "":
                results.append("自发性眼震速度：" + spontaneous.get("自发性眼震速度", ""))
            if spontaneous.get("自发性眼震固视抑制", "") != "":
                results.append("自发性眼震固视抑制：" + spontaneous.get("自发性眼震固视抑制", ""))
        
        # 凝视性眼震
        gaze = data.get("凝视性眼震", {})
        if not is_dict_empty(gaze):
            if gaze.get("凝视性眼震模式（上）", "") != "":
                results.append("凝视性眼震模式（上）: " + gaze.get("凝视性眼震模式（上）", ""))
            if gaze.get("凝视性眼震速度（上）", "") != "":
                results.append("凝视性眼震速度（上，度/秒）: " + gaze.get("凝视性眼震速度（上）", ""))
            if gaze.get("凝视性眼震模式（下）", "") != "":
                results.append("凝视性眼震模式（下）: " + gaze.get("凝视性眼震模式（下）", ""))
            if gaze.get("凝视性眼震速度（下）", "") != "":
                results.append("凝视性眼震速度（下，度/秒）: " + gaze.get("凝视性眼震速度（下）", ""))
            if gaze.get("凝视性眼震模式（左）", "") != "":
                results.append("凝视性眼震模式（左）: " + gaze.get("凝视性眼震模式（左）", ""))
            if gaze.get("凝视性眼震速度（右）", "") != "":
                results.append("凝视性眼震速度（右，度/秒）: " + gaze.get("凝视性眼震速度（右）", ""))
        
        # 头脉冲试验
        hit = data.get("头脉冲试验", {})
        if not is_dict_empty(hit):
            if hit.get("VOR增益（左外半规管）", "") != "":
                results.append("VOR增益（左外半规管）：" + hit.get("VOR增益（左外半规管）", ""))
            if hit.get("PR分数（左外半规管）", "") != "":
                results.append("PR分数（左外半规管，%）：" + hit.get("PR分数（左外半规管）", ""))
            if hit.get("VOR增益（右外半规管）", "") != "":
                results.append("VOR增益（右外半规管）：" + hit.get("VOR增益（右外半规管）", ""))
            if hit.get("PR分数（右外半规管）", "") != "":
                results.append("PR分数（右外半规管，%）：" + hit.get("PR分数（右外半规管）", ""))
            if hit.get("VOR增益（左前半规管）", "") != "":
                results.append("VOR增益（左前半规管）：" + hit.get("VOR增益（左前半规管）", ""))
            if hit.get("PR分数（右前半规管）", "") != "":
                results.append("PR分数（右前半规管，%）：" + hit.get("PR分数（右前半规管）", ""))
            if hit.get("VOR增益（左后半规管）", "") != "":
                results.append("VOR增益（左后半规管）：" + hit.get("VOR增益（左后半规管）", ""))
            if hit.get("PR分数（左后半规管）", "") != "":
                results.append("PR分数（左后半规管，%）：" + hit.get("PR分数（左后半规管）", ""))
            if hit.get("VOR增益（右后半规管）", "") != "":
                results.append("VOR增益（右后半规管）：" + hit.get("VOR增益（右后半规管）", ""))
            if hit.get("PR分数（右后半规管）", "") != "":
                results.append("PR分数（右后半规管，%）：" + hit.get("PR分数（右后半规管）", ""))
        
            # 补偿性扫视波（多选项）
            compensatory_waves = hit.get("头脉冲试验补偿性扫视波", [])
            if compensatory_waves != []:
                if isinstance(compensatory_waves, list):
                    results.append("头脉冲试验补偿性扫视波: " + "、".join(compensatory_waves))
                else:
                    results.append("头脉冲试验补偿性扫视波: " + str(compensatory_waves))
        
        # 头脉冲抑制试验
        hit_suppression = data.get("头脉冲抑制试验", {})
        if not is_dict_empty(hit_suppression):
            if hit_suppression.get("头脉冲抑制试验增益（左外半规管）", "") != "":
                results.append("头脉冲抑制试验增益（左外半规管）：" + hit_suppression.get("头脉冲抑制试验增益 (左外半规管)", ""))
            if hit_suppression.get("头脉冲抑制试验增益（右外半规管）", "") != "":
                results.append("头脉冲抑制试验增益（右外半规管）： " + hit_suppression.get("头脉冲抑制试验增益 (右外半规管)", ""))
            
            # 补偿性扫视波（多选项）
            compensation_waves = hit_suppression.get("头脉冲抑制试验补偿性扫视波", [])
            if compensation_waves != []:
                if isinstance(compensation_waves, list):
                    results.append("头脉冲抑制试验补偿性扫视波：" + "、".join(compensation_waves))
                else:
                    results.append("头脉冲抑制试验补偿性扫视波：" + str(compensation_waves))
        
        # 眼位反向偏斜
        skew = data.get("眼位反向偏斜", {})
        if not is_dict_empty(skew):
            if skew.get("眼位反向偏斜（HR, 度）", "") != "":
                results.append("眼位反向偏斜（HR, 度）：" + skew.get("眼位反向偏斜 (HR, 度)", ""))
            if skew.get("眼位反向偏斜（VR, 度）", "") != "":
                results.append("眼位反向偏斜（VR, 度）：" + skew.get("眼位反向偏斜 (VR, 度)", ""))
        
        # 扫视检查
        saccade = data.get("扫视检查", {})
        if not is_dict_empty(saccade):
            if saccade.get("扫视延迟时间（右向, 毫秒）", "") != "":
                results.append("扫视延迟时间（右向, 毫秒）：" + saccade.get("扫视延迟时间 (右向, 毫秒)", ""))
            if saccade.get("扫视延迟时间（左向, 毫秒）", "") != "":
                results.append("扫视延迟时间（左向, 毫秒）：" + saccade.get("扫视延迟时间 (左向, 毫秒)", ""))
            if saccade.get("扫视峰速度（右向, 度/秒）", "") != "":
                results.append("扫视峰速度（右向, 度/秒）：" + saccade.get("扫视峰速度 (右向, 度/秒)", ""))
            if saccade.get("扫视峰速度（左向, 度/秒）", "") != "":
                results.append("扫视峰速度（左向, 度/秒）：" + saccade.get("扫视峰速度 (左向, 度/秒)", ""))
            if saccade.get("扫视精确度（右向, %）", "") != "":
                results.append("扫视精确度（右向, %）：" + saccade.get("扫视精确度 (右向, %)", ""))
            if saccade.get("扫视精确度（左向, %）", "") != "":
                results.append("扫视精确度（左向, %）：" + saccade.get("扫视精确度 (左向, %)", ""))
        
        # 视觉增强前庭-眼反射试验
        vevor = data.get("视觉增强前庭-眼反射试验", {})
        if not is_dict_empty(vevor):
            if vevor.get("检查结果", "") != "":
                results.append("视觉增强前庭-眼反射试验检查结果：" + vevor.get("检查结果", ""))
        
        # 前庭-眼反射抑制试验
        vor_suppression = data.get("前庭-眼反射抑制试验", {})
        if not is_dict_empty(vor_suppression):
            if vor_suppression.get("检查结果", "") != "":
                results.append("前庭-眼反射抑制试验检查结果：" + vor_suppression.get("检查结果", ""))
        
        # 摇头试验
        hst = data.get("摇头试验", {})
        if not is_dict_empty(hst):
            if hst.get("眼震模式", "") != "":
                results.append("眼震模式（摇头试验）：" + hst.get("眼震模式", ""))
            if hst.get("眼震速度", "") != "":
                results.append("眼震速度（摇头试验，度/秒）：" + hst.get("眼震速度", ""))
            if hst.get("摇头方向", "") != "":
                results.append("摇头诱发眼震方向： " + hst.get("摇头方向", ""))
        
        # Dix-Hallpike试验
        dix = data.get("位置试验 (Dix-Hallpike试验)", {})
        if not is_dict_empty(dix):
            if dix.get("右侧眼震模式", "") != "":   
                results.append("眼震模式（右侧D-H试验）：" + dix.get("右侧眼震模式", ""))
            if dix.get("右侧坐起眼震模式", "") != "":
                results.append("坐起眼震模式（右侧D-H试验）：" + dix.get("右侧坐起眼震模式", ""))
            if dix.get("右侧出现眩晕/头晕", "") != "":
                results.append("出现眩晕/头晕（右侧D-H试验）：" + dix.get("右侧出现眩晕/头晕", ""))
            if dix.get("右侧眼震潜伏期 (秒)", "") != "":
                results.append("眼震潜伏期（右侧D-H试验，秒）：" + dix.get("右侧眼震潜伏期 (秒)", ""))
            if dix.get("右侧眼震持续时长 (秒)", "") != "":
                results.append("眼震持续时长（右侧D-H试验，秒）：" + dix.get("右侧眼震持续时长 (秒)", ""))
            if dix.get("右侧眼震最大速度 (度/秒)", "") != "":
                results.append("眼震最大速度（右侧D-H试验，度/秒）：" + dix.get("右侧眼震最大速度 (度/秒)", ""))
            if dix.get("右侧眼震疲劳性", "") != "": 
                results.append("眼震疲劳性（右侧D-H试验）：" + dix.get("右侧眼震疲劳性", ""))
            if dix.get("左侧眼震模式", "") != "":
                results.append("眼震模式（左侧D-H试验）：" + dix.get("左侧眼震模式", ""))
            if dix.get("左侧坐起眼震模式", "") != "":
                results.append("坐起眼震模式（左侧D-H试验）：" + dix.get("左侧坐起眼震模式", ""))
            if dix.get("左侧出现眩晕/头晕", "") != "":
                results.append("出现眩晕/头晕（左侧D-H试验）：" + dix.get("左侧出现眩晕/头晕", ""))
            if dix.get("左侧眼震潜伏期 (秒)", "") != "":
                results.append("眼震潜伏期（左侧D-H试验，秒）：" + dix.get("左侧眼震潜伏期 (秒)", ""))
            if dix.get("左侧眼震持续时长 (秒)", "") != "":
                results.append("眼震持续时长（左侧D-H试验，秒）：" + dix.get("左侧眼震持续时长 (秒)", ""))
            if dix.get("左侧眼震最大速度 (度/秒)", "") != "":
                results.append("眼震最大速度（左侧D-H试验，度/秒）：" + dix.get("左侧眼震最大速度 (度/秒)", ""))
            if dix.get("左侧眼震疲劳性", "") != "":
                results.append("眼震疲劳性（左侧D-H试验）：" + dix.get("左侧眼震疲劳性", ""))
                
        # 仰卧滚转试验
        roll = data.get("位置试验 (仰卧滚转试验)", {})
        if not is_dict_empty(roll):
            if roll.get("右侧眼震模式", "") != "":  
                results.append("眼震模式（右侧仰卧滚转试验）：" + roll.get("右侧眼震模式", ""))
            if roll.get("右侧出现眩晕/头晕", "") != "":
                results.append("出现眩晕/头晕（右侧仰卧滚转试验）：" + roll.get("右侧出现眩晕/头晕", ""))
            if roll.get("右侧眼震潜伏期 (秒)", "") != "":
                results.append("眼震潜伏期（右侧仰卧滚转试验，秒）：" + roll.get("右侧眼震潜伏期 (秒)", ""))
            if roll.get("右侧眼震持续时长 (秒)", "") != "":
                results.append("眼震持续时长（右侧仰卧滚转试验，秒）：" + roll.get("右侧眼震持续时长 (秒)", ""))
            if roll.get("右侧眼震最大速度 (度/秒)", "") != "":
                results.append("眼震最大速度（右侧仰卧滚转试验，度/秒）：" + roll.get("右侧眼震最大速度 (度/秒)", ""))
            if roll.get("左侧眼震模式", "") != "":
                results.append("眼震模式（左侧仰卧滚转试验）：" + roll.get("左侧眼震模式", ""))
            if roll.get("左侧出现眩晕/头晕", "") != "":
                results.append("出现眩晕/头晕（左侧仰卧滚转试验）：" + roll.get("左侧出现眩晕/头晕", ""))
            if roll.get("左侧眼震潜伏期 (秒)", "") != "":
                results.append("眼震潜伏期（左侧仰卧滚转试验，秒）：" + roll.get("左侧眼震潜伏期 (秒)", ""))
            if roll.get("左侧眼震持续时长 (秒)", "") != "":
                results.append("眼震持续时长（左侧仰卧滚转试验，秒）：" + roll.get("左侧眼震持续时长 (秒)", ""))
            if roll.get("左侧眼震最大速度 (度/秒)", "") != "":
                results.append("眼震最大速度（左侧仰卧滚转试验，度/秒）：" + roll.get("左侧眼震最大速度 (度/秒)", ""))
    
        # 其他位置试验
        other = data.get("位置试验（其他）", {})
        if not is_dict_empty(other):
            if other.get("坐位-平卧试验", "") != "":    
                results.append("坐位-平卧试验：" + other.get("坐位-平卧试验", ""))
            if other.get("坐位-低头试验", "") != "":
                results.append("坐位-低头试验：" + other.get("坐位-低头试验", ""))
            if other.get("坐位-仰头试验", "") != "":
                results.append("坐位-仰头试验：" + other.get("坐位-仰头试验", ""))
            if other.get("零平面", "") != "":
                results.append("零平面: " + other.get("零平面", ""))
        
        # 视跟踪
        tracking = data.get("视跟踪", {})
        if not is_dict_empty(tracking):
            if tracking.get("视跟踪曲线分型", "") != "":    
                results.append("视跟踪曲线分型: " + tracking.get("视跟踪曲线分型", ""))
            if tracking.get("视跟踪增益", "") != "":
                results.append("视跟踪增益: " + tracking.get("视跟踪增益", ""))
        
        # 视动性眼震
        okn = data.get("视动性眼震", {})
        if not is_dict_empty(okn):
            if okn.get("水平视标不对称性 (%)", "") != "":
                results.append("视动性眼震（水平视靶）不对称性 (%): " + okn.get("水平视标不对称性 (%)", ""))
            if okn.get("向右视标增益", "") != "":
                results.append("视动性眼震增益（向右视靶）: " + okn.get("向右视标增益", ""))
            if okn.get("向左视标增益", "") != "":
                results.append("视动性眼震增益（向左视靶）: " + okn.get("向左视标增益", ""))
            if okn.get("垂直视标不对称性 (%)", "") != "":
                results.append("视动性眼震（垂直视靶）不对称性 (%): " + okn.get("垂直视标不对称性 (%)", ""))
            if okn.get("向上视标增益", "") != "":
                results.append("视动性眼震增益（向上视靶）: " + okn.get("向上视标增益", ""))
            if okn.get("向下视标增益", "") != "":
                results.append("视动性眼震增益（向下视靶）: " + okn.get("向下视标增益", ""))
            if okn.get("检查结果", "") != "":
                results.append("视动性眼震检查结果: " + okn.get("检查结果", ""))
    
        # 瘘管试验
        fistula = data.get("瘘管试验", {})
        if not is_dict_empty(fistula):
            fistula_results = fistula.get("瘘管试验", [])
            if fistula_results != []:
                if isinstance(fistula_results, list):
                    results.append("瘘管试验：" + "、".join(fistula_results))
                else:
                    results.append("瘘管试验：" + str(fistula_results))
            
        
        # 温度试验
        caloric = data.get("温度试验", {})
        if not is_dict_empty(caloric):
            if caloric.get("单侧减弱侧别 (UW)", "") != "":  
                results.append("单侧减弱侧别 (UW): " + caloric.get("单侧减弱侧别 (UW)", ""))
            if caloric.get("单侧减弱数值 (UW, %)", "") != "":
                results.append("单侧减弱数值 (UW, %): " + caloric.get("单侧减弱数值 (UW, %)", ""))
            if caloric.get("优势偏向侧别 (DP)", "") != "":
                results.append("优势偏向侧别 (DP, 度/秒): " + caloric.get("优势偏向侧别 (DP)", ""))
            if caloric.get("优势偏向数值 (DP, 度/秒)", "") != "":
                results.append("优势偏向数值 (DP, 度/秒): " + caloric.get("优势偏向数值 (DP, 度/秒)", ""))
            if caloric.get("最大慢相速度总和（右耳，度/秒）", "") != "":
                results.append("最大慢相速度总和（右耳，度/秒）: " + caloric.get("最大慢相速度总和（右耳, 度/秒）", ""))
            if caloric.get("最大慢相速度总和（左耳，度/秒）", "") != "":
                results.append("最大慢相速度总和（左耳，度/秒）: " + caloric.get("最大慢相速度总和（左耳, 度/秒）", ""))
            if caloric.get("固视抑制指数 (FI, %)", "") != "":
                results.append("固视抑制指数 (FI, %): " + caloric.get("固视抑制指数 (FI, %)", ""))
        
        # cVEMP
        cvemp = data.get("颈肌前庭诱发肌源性电位", {})
        if not is_dict_empty(cvemp):
            
            if cvemp.get("右耳声强阈值 (分贝)", "") != "":
                results.append("声强阈值（cVEMP，右耳，分贝）: " + cvemp.get("右耳声强阈值 (分贝)", ""))
            if cvemp.get("右耳P13波潜伏期 (毫秒)", "") != "":
                results.append("P13波潜伏期（右耳，毫秒）: " + cvemp.get("右耳P13波潜伏期 (毫秒)", ""))
            if cvemp.get("右耳N23波潜伏期 (毫秒)", "") != "":
                results.append("N23波潜伏期（右耳，毫秒）: " + cvemp.get("右耳N23波潜伏期 (毫秒)", ""))
            if cvemp.get("右耳P13-N23波间期 (毫秒)", "") != "":
                results.append("P13-N23波间期（右耳，毫秒）: " + cvemp.get("右耳P13-N23波间期 (毫秒)", ""))
            if cvemp.get("右耳P13波振幅 (微伏)", "") != "":
                results.append("P13波振幅（右耳，微伏）: " + cvemp.get("右耳P13波振幅 (微伏)", ""))
            if cvemp.get("右耳P13-N23波振幅 (微伏)", "") != "":
                results.append("P13-N23波振幅（右耳，微伏）: " + cvemp.get("右耳P13-N23波振幅 (微伏)", ""))
            if cvemp.get("左耳声强阈值 (分贝)", "") != "":
                results.append("声强阈值（cVEMP，左耳，分贝）: " + cvemp.get("左耳声强阈值 (分贝)", ""))
            if cvemp.get("左耳P13波潜伏期 (毫秒)", "") != "":
                results.append("P13波潜伏期（左耳，毫秒）: " + cvemp.get("左耳P13波潜伏期 (毫秒)", ""))
            if cvemp.get("左耳N23波潜伏期 (毫秒)", "") != "":
                results.append("N23波潜伏期（左耳，毫秒）: " + cvemp.get("左耳N23波潜伏期 (毫秒)", ""))
            if cvemp.get("左耳P13-N23波间期 (毫秒)", "") != "":
                results.append("P13-N23波间期（左耳，毫秒）: " + cvemp.get("左耳P13-N23波间期 (毫秒)", ""))
            if cvemp.get("左耳P13波振幅 (微伏)", "") != "":
                results.append("P13波振幅（左耳，微伏）: " + cvemp.get("左耳P13波振幅 (微伏)", ""))
            if cvemp.get("左耳N23波振幅 (微伏)", "") != "":
                results.append("N23波振幅（左耳，微伏）: " + cvemp.get("左耳N23波振幅 (微伏)", ""))
            if cvemp.get("左耳P13-N23波振幅 (微伏)", "") != "":
                results.append("P13-N23波振幅（左耳，微伏）: " + cvemp.get("左耳P13-N23波振幅 (微伏)", ""))
            if cvemp.get("cVEMP耳间不对称比 (%)", "") != "":
                results.append("cVEMP耳间不对称比 (%): " + cvemp.get("cVEMP耳间不对称比 (%)", ""))
        
        # oVEMP
        ovemp = data.get("眼肌前庭诱发肌源性电位 (oVEMP)", {})
        if not is_dict_empty(ovemp):
            if ovemp.get("右耳声强阈值 (分贝)", "") != "":
                results.append("声强阈值（oVEMP，右耳，分贝）: " + ovemp.get("右耳声强阈值 (分贝)", ""))
            if ovemp.get("右耳N10波潜伏期 (毫秒)", "") != "":
                results.append("N10波潜伏期（右耳，毫秒）: " + ovemp.get("右耳N10波潜伏期 (毫秒)", ""))
            if ovemp.get("右耳P15波潜伏期 (毫秒)", "") != "":
                results.append("P15波潜伏期（右耳，毫秒）: " + ovemp.get("右耳P15波潜伏期 (毫秒)", ""))
            if ovemp.get("右耳N10-P15波间期 (毫秒)", "") != "":
                results.append("N10-P15波间期（右耳，毫秒）: " + ovemp.get("右耳N10-P15波间期 (毫秒)", ""))
            if ovemp.get("右耳N10波振幅 (微伏)", "") != "":
                results.append("N10波振幅（右耳，微伏）: " + ovemp.get("右耳N10波振幅 (微伏)", ""))
            if ovemp.get("右耳N10-P15波振幅 (微伏)", "") != "":
                results.append("N10-P15波振幅（右耳，微伏）: " + ovemp.get("右耳N10-P15波振幅 (微伏)", ""))
            if ovemp.get("左耳声强阈值 (分贝)", "") != "":
                results.append("声强阈值（oVEMP，左耳，分贝）: " + ovemp.get("左耳声强阈值 (分贝)", ""))
            if ovemp.get("左耳N10波潜伏期 (毫秒)", "") != "":
                results.append("N10波潜伏期（左耳，毫秒）: " + ovemp.get("左耳N10波潜伏期 (毫秒)", ""))
            if ovemp.get("左耳P15波潜伏期 (毫秒)", "") != "":
                results.append("P15波潜伏期（左耳，毫秒）: " + ovemp.get("左耳P15波潜伏期 (毫秒)", ""))
            if ovemp.get("左耳N10-P15波间期 (毫秒)", "") != "":
                results.append("N10-P15波间期（左耳，毫秒）: " + ovemp.get("左耳N10-P15波间期 (毫秒)", ""))
            if ovemp.get("左耳N10波振幅 (微伏)", "") != "":
                results.append("N10波振幅（左耳，微伏）: " + ovemp.get("左耳N10波振幅 (微伏)", ""))
            if ovemp.get("左耳P15波振幅 (微伏)", "") != "":
                results.append("P15波振幅（左耳，微伏）: " + ovemp.get("左耳P15波振幅 (微伏)", ""))
            if ovemp.get("左耳N10-P15波振幅 (微伏)", "") != "":
                results.append("N10-P15波振幅（左耳，微伏）: " + ovemp.get("左耳N10-P15波振幅 (微伏)", ""))
            if ovemp.get("oVEMP耳间不对称性 (%)", "") != "":
                results.append("oVEMP耳间不对称性 (%): " + ovemp.get("oVEMP耳间不对称性 (%)", ""))
        
        # SVV
        svv = data.get("主观视觉垂直线 (SVV)", {})
        if not is_dict_empty(svv):
            if svv.get("偏斜方向", "") != "":
                results.append("主观视觉垂直线偏斜方向: " + svv.get("偏斜方向", ""))
            if svv.get("偏斜角度（度）", "") != "":
                results.append("主观视觉垂直线偏斜角度（度）: " + svv.get("偏斜角度（度）", ""))
            
            
        # 结果中间用;隔开
        results = ';'.join(results)

        return results

    def load_config(self):
        config = json.load(open(self.config_file, 'r'))
        self.font_name = config['font_name']
        self.font_path = config['font_path']
    
    def get_data(self):
        return {}



