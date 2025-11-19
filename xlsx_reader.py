#!/usr/bin/env python3
"""
Extract the structural layout of an Excel workbook into a JSON description.

Usage:
    python xlsx_reader.py --input path/to/workbook.xlsx --output layout.json
"""

from __future__ import annotations

import argparse
import json
import os
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any, Dict, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Fill, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="读取 Excel 模板的单元格布局并导出为 JSON 文件。"
    )
    parser.add_argument(
        "--input",
        "-i",
        required=True,
        help="需要解析的 Excel xlsx 文件路径。",
    )
    parser.add_argument(
        "--output",
        "-o",
        help="输出的 JSON 文件路径，默认为同名 layout.json。",
    )
    parser.add_argument(
        "--indent",
        type=int,
        default=2,
        help="JSON 缩进空格数，默认 2。",
    )
    return parser.parse_args()


def serialize_scalar(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def json_default(value: Any) -> Any:
    serialized = serialize_scalar(value)
    if serialized is not value:
        return serialized
    if hasattr(value, "value"):
        try:
            return serialize_scalar(value.value)
        except Exception:
            return value.value
    return str(value)


def color_to_dict(color) -> Optional[Dict[str, Any]]:
    if color is None:
        return None
    attrs = {}
    for attr in ("type", "theme", "tint", "indexed", "auto"):
        val = getattr(color, attr, None)
        if val not in (None, False):
            attrs[attr] = val
    rgb = getattr(color, "rgb", None)
    if rgb:
        attrs["rgb"] = rgb
    if not attrs:
        return None
    return attrs


def font_to_dict(font: Font) -> Optional[Dict[str, Any]]:
    if font is None:
        return None
    data = {
        "name": font.name,
        "size": font.size,
        "bold": font.bold,
        "italic": font.italic,
        "underline": font.underline if font.underline else None,
        "strike": font.strike,
        "color": color_to_dict(font.color),
    }
    return {k: v for k, v in data.items() if v not in (None, False)}


def alignment_to_dict(alignment: Alignment) -> Optional[Dict[str, Any]]:
    if alignment is None:
        return None
    data = {
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
        "wrap_text": alignment.wrap_text,
        "shrink_to_fit": alignment.shrink_to_fit,
        "text_rotation": alignment.text_rotation,
    }
    return {k: v for k, v in data.items() if v not in (None, False)}


def side_to_dict(side: Side) -> Optional[Dict[str, Any]]:
    if side is None:
        return None
    data = {
        "style": side.style,
        "color": color_to_dict(side.color),
    }
    return {k: v for k, v in data.items() if v is not None}


def border_to_dict(border: Border) -> Optional[Dict[str, Any]]:
    if border is None:
        return None
    data = {
        "left": side_to_dict(border.left),
        "right": side_to_dict(border.right),
        "top": side_to_dict(border.top),
        "bottom": side_to_dict(border.bottom),
        "diagonal": side_to_dict(border.diagonal),
        "diagonal_direction": border.diagonal_direction,
    }
    cleaned = {k: v for k, v in data.items() if v is not None}
    return cleaned or None


def fill_to_dict(fill: Fill) -> Optional[Dict[str, Any]]:
    if fill is None or not isinstance(fill, PatternFill):
        return None
    data = {
        "pattern_type": fill.patternType,
        "fg_color": color_to_dict(fill.fgColor),
        "bg_color": color_to_dict(fill.bgColor),
    }
    return {k: v for k, v in data.items() if v is not None}


def cell_to_dict(cell: Cell, value_cell: Cell, merged_map: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    column_index = getattr(cell, "col_idx", None)
    row_index = getattr(cell, "row", None)
    if column_index is None or row_index is None:
        column_letter, row_val = coordinate_from_string(cell.coordinate)
        column_index = column_index_from_string(column_letter)
        row_index = row_val

    info: Dict[str, Any] = {
        "coordinate": cell.coordinate,
        "row": row_index,
        "column": column_index,
        "data_type": cell.data_type,
    }

    if cell.data_type == "f":
        info["formula"] = cell.value
        info["value"] = serialize_scalar(value_cell.value)
    else:
        info["value"] = serialize_scalar(cell.value)

    if cell.has_style:
        info["number_format"] = cell.number_format

        font_dict = font_to_dict(cell.font)
        if font_dict:
            info["font"] = font_dict

        alignment_dict = alignment_to_dict(cell.alignment)
        if alignment_dict:
            info["alignment"] = alignment_dict

        border_dict = border_to_dict(cell.border)
        if border_dict:
            info["border"] = border_dict

        fill_dict = fill_to_dict(cell.fill)
        if fill_dict:
            info["fill"] = fill_dict

        if cell.protection:
            locked = cell.protection.locked
            hidden = cell.protection.hidden
            if locked is False or hidden is True:
                info["protection"] = {
                    "locked": locked,
                    "hidden": hidden,
                }

    if cell.hyperlink:
        info["hyperlink"] = cell.hyperlink.target

    if cell.comment:
        info["comment"] = cell.comment.text

    merged_info = merged_map.get(cell.coordinate)
    if merged_info:
        info["merged"] = merged_info

    return {k: v for k, v in info.items() if v not in (None, [], {})}


def build_merged_map(sheet) -> Dict[str, Dict[str, Any]]:
    merged_map: Dict[str, Dict[str, Any]] = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        master_coord = f"{get_column_letter(min_col)}{min_row}"
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                coord = f"{get_column_letter(col_idx)}{row_idx}"
                merged_map[coord] = {
                    "range": str(merged_range),
                    "is_master": coord == master_coord,
                }
    return merged_map


def column_dimensions_to_dict(sheet) -> Dict[str, Dict[str, Any]]:
    result: Dict[str, Dict[str, Any]] = {}
    for column_letter, dim in sheet.column_dimensions.items():
        data: Dict[str, Any] = {}
        if dim.width is not None:
            data["width"] = dim.width
        if dim.hidden:
            data["hidden"] = True
        if dim.outlineLevel:
            data["outline_level"] = dim.outlineLevel
        if dim.bestFit:
            data["best_fit"] = True
        if dim.customWidth:
            data["custom_width"] = True
        if data:
            result[column_letter] = data
    return result


def row_dimensions_to_dict(sheet) -> Dict[str, Dict[str, Any]]:
    result: Dict[str, Dict[str, Any]] = {}
    for row_idx, dim in sheet.row_dimensions.items():
        data: Dict[str, Any] = {}
        if dim.height is not None:
            data["height"] = dim.height
        if dim.hidden:
            data["hidden"] = True
        if dim.outlineLevel:
            data["outline_level"] = dim.outlineLevel
        if dim.customHeight:
            data["custom_height"] = True
        if data:
            result[str(row_idx)] = data
    return result


def sheet_to_dict(sheet, value_sheet) -> Dict[str, Any]:
    merged_map = build_merged_map(sheet)
    sheet_info: Dict[str, Any] = {
        "name": sheet.title,
        "dimension": sheet.calculate_dimension(),
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
    }

    if sheet.freeze_panes:
        sheet_info["freeze_panes"] = sheet.freeze_panes.coordinate if isinstance(sheet.freeze_panes, Cell) else sheet.freeze_panes

    merged_ranges = [str(rng) for rng in sheet.merged_cells.ranges]
    if merged_ranges:
        sheet_info["merged_cells"] = merged_ranges

    column_dims = column_dimensions_to_dict(sheet)
    if column_dims:
        sheet_info["column_dimensions"] = column_dims

    row_dims = row_dimensions_to_dict(sheet)
    if row_dims:
        sheet_info["row_dimensions"] = row_dims

    cells = []
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            # 跳过完全空白且无样式的单元格
            if (
                cell.value is None
                and cell.data_type != "f"
                and not cell.has_style
                and cell.coordinate not in merged_map
                and cell.hyperlink is None
                and cell.comment is None
            ):
                continue
            value_cell = value_sheet[cell.coordinate]
            cells.append(cell_to_dict(cell, value_cell, merged_map))

    sheet_info["cells"] = cells
    return sheet_info


def workbook_to_dict(path: str) -> Dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    workbook_values = load_workbook(path, data_only=True)

    wb_info: Dict[str, Any] = {
        "file": os.path.basename(path),
        "sheet_names": workbook.sheetnames,
        "sheets": [],
    }

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        value_sheet = workbook_values[sheet_name]
        wb_info["sheets"].append(sheet_to_dict(sheet, value_sheet))

    return wb_info


def main() -> None:
    args = parse_args()
    input_path = os.path.abspath(args.input)

    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"未找到指定的 Excel 文件: {input_path}")

    output_path = (
        os.path.abspath(args.output)
        if args.output
        else os.path.splitext(input_path)[0] + "_layout.json"
    )

    workbook_info = workbook_to_dict(input_path)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(
            workbook_info,
            f,
            ensure_ascii=False,
            indent=args.indent,
            default=json_default,
        )

    print(f"已生成布局描述文件: {output_path}")


if __name__ == "__main__":
    main()

