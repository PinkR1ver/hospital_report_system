"""
Excel报告生成器 - 基于JSON配置文件的通用Excel生成系统
支持动态布局、条件显示、样式定义、数据格式化等功能
"""
import json
import os
from datetime import datetime
import re

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """基于配置文件的Excel报告生成器"""

    def __init__(self, config_path=None, database_path="vest_database", template_id=None):
        """初始化生成器

        Args:
            config_path: 直接指定配置文件路径（优先）
            database_path: 数据库路径
            template_id: 模板ID（从templates_index.json中查找）
        """
        self.database_path = database_path
        self.template_id = template_id

        if config_path:
            self.config = self._load_config(config_path)
        else:
            if template_id is None:
                index_path = os.path.join(database_path, "templates", "templates_index.json")
                if os.path.exists(index_path):
                    with open(index_path, "r", encoding="utf-8") as f:
                        index_data = json.load(f)
                    template_id = index_data.get("default_template", "default_template")
                else:
                    template_id = "default_template"
                self.template_id = template_id

            config_path = self._find_template_config(database_path, self.template_id)
            self.config = self._load_config(config_path)

        self.template_file = self.config.get("template_file")
        self.styles = self._build_styles(self.config.get("styles", {}))
        self.current_row_tracker = {}  # 跟踪每个工作表的当前行

    def _find_template_config(self, database_path, template_id):
        """从模板索引中查找配置文件路径"""
        index_path = os.path.join(database_path, "templates", "templates_index.json")
        if not os.path.exists(index_path):
            # 如果没有索引文件，尝试直接加载默认配置
            default_path = os.path.join(database_path, "templates", "default_template_config.json")
            if os.path.exists(default_path):
                return default_path
            raise FileNotFoundError(f"模板索引文件不存在: {index_path}")
        with open(index_path, 'r', encoding='utf-8') as f:
            index_data = json.load(f)

        templates = index_data.get("templates", [])
        for template in templates:
            if template.get("id") == template_id:
                config_file = template.get("config_file")
                if config_file:
                    config_path = os.path.join(database_path, "templates", config_file)
                    if os.path.exists(config_path):
                        return config_path
                    else:
                        raise FileNotFoundError(f"模板配置文件不存在: {config_path}")
        
        raise ValueError(f"未找到模板ID: {template_id}")
    
    def _load_config(self, config_path):
        """加载配置文件"""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)

            # 如果配置中的template_file是相对路径，尝试从数据库路径解析
            template_file = config.get("template_file", "")
            if template_file and not os.path.isabs(template_file):
                # 尝试多个可能的路径
                possible_paths = [
                    template_file,  # 原始路径
                    os.path.join(self.database_path, "templates", template_file),  # 数据库templates目录
                    os.path.join(os.path.dirname(config_path), template_file),  # 配置文件同级目录
                    template_file  # 项目根目录
                ]
                for path in possible_paths:
                    if os.path.exists(path):
                        config["template_file"] = path
                        break
            
            return config
        except FileNotFoundError:
            raise FileNotFoundError(f"配置文件不存在: {config_path}")
        except json.JSONDecodeError as e:
            raise ValueError(f"配置文件格式错误: {e}")
    
    def _build_styles(self, styles_config):
        """构建样式对象"""
        built_styles = {}
        for style_name, style_config in styles_config.items():
            # 构建字体
            font_config = style_config.get("font", {})
            font = Font(
                name=font_config.get("name", "宋体"),
                size=font_config.get("size", 11),
                bold=font_config.get("bold", False)
            )
            
            # 构建填充
            fill = None
            fill_config = style_config.get("fill", {})
            if fill_config.get("type") == "solid":
                fill = PatternFill(
                    start_color=fill_config.get("color", "FFFFFF"),
                    end_color=fill_config.get("color", "FFFFFF"),
                    fill_type="solid"
                )
            
            # 构建对齐
            align_config = style_config.get("alignment", {})
            alignment = Alignment(
                horizontal=align_config.get("horizontal", "center"),
                vertical=align_config.get("vertical", "center"),
                wrap_text=align_config.get("wrap_text", False),
                shrink_to_fit=align_config.get("shrink_to_fit", False),
            )
            
            # 构建边框
            border_config = style_config.get("border", {})
            border_side = Side(
                style=border_config.get("all", "thin"),
                color=border_config.get("color", "000000")
            )
            border = Border(
                left=border_side,
                right=border_side,
                top=border_side,
                bottom=border_side
            )
            
            built_styles[style_name] = {
                "font": font,
                "fill": fill,
                "alignment": alignment,
                "border": border
            }

        return built_styles
    
    def _is_empty(self, value):
        """检查值是否为空"""
        empty_values = self.config.get("empty_value_handling", {}).get("check_empty_values", 
            ["", None, [], {}, "未知", "无", "N/A", "NULL"])
        return value in empty_values or (isinstance(value, (list, dict)) and len(value) == 0)
    
    def _is_section_empty(self, data, section_key):
        """检查整个section是否为空"""
        section_data = data.get(section_key, {})
        if not isinstance(section_data, dict):
            return self._is_empty(section_data)
        
        # 检查字典中所有值是否都为空
        for value in section_data.values():
            if not self._is_empty(value):
                return False
        return True
    
    def _get_data_value(self, data, data_path):
        """根据数据路径获取值（支持嵌套路径，如 '基本信息.ID'）"""
        keys = data_path.split('.')
        value = data
        for key in keys:
            if isinstance(value, dict):
                value = value.get(key, "")
            else:
                return ""
        return value
    
    def _format_value(self, value, format_type="text"):
        """格式化值"""
        formatters = self.config.get("data_formatters", {})
        
        if format_type == "date":
            date_config = formatters.get("date", {})
            input_fmt = date_config.get("input_format", "%Y/%m/%d")
            output_fmt = date_config.get("output_format", "%Y/%m/%d")
            try:
                if isinstance(value, str):
                    dt = datetime.strptime(value, input_fmt)
                    return dt.strftime(output_fmt)
            except:
                return value
        
        elif format_type == "number":
            num_config = formatters.get("number", {})
            try:
                num = float(value)
                places = num_config.get("decimal_places", 2)
                return round(num, places)
            except:
                return value
        
        elif format_type == "list":
            list_config = formatters.get("list", {})
            if isinstance(value, list):
                if list_config.get("join", True):
                    separator = list_config.get("separator", "、")
                    return separator.join(str(v) for v in value if v)
                return value
            return value
        
        return value
    
    def _apply_style(self, cell, style_name):
        """应用样式到单元格"""
        if style_name not in self.styles:
            return
        
        style = self.styles[style_name]
        cell.font = style["font"]
        if style["fill"]:
            cell.fill = style["fill"]
        cell.alignment = style["alignment"]
        cell.border = style["border"]
    
    def _resolve_cell_range(self, range_str, current_row):
        """解析单元格范围，支持 {row} 占位符"""
        # 替换 {row} 为当前行号
        range_str = range_str.replace("{row}", str(current_row))
        
        # 处理 {row+1}, {row+2} 等
        def replace_row_offset(match):
            offset = int(match.group(1) or 0)
            return str(current_row + offset)
        
        range_str = re.sub(r'\{row([+-]\d+)?\}', replace_row_offset, range_str)
        
        return range_str
    
    def _get_current_row(self, worksheet_name):
        """获取工作表的当前行"""
        return self.current_row_tracker.get(worksheet_name, 1)
    
    def _set_current_row(self, worksheet_name, row):
        """设置工作表的当前行"""
        self.current_row_tracker[worksheet_name] = row
    
    def _increment_row(self, worksheet_name, increment=1):
        """增加工作表的当前行"""
        current = self._get_current_row(worksheet_name)
        self._set_current_row(worksheet_name, current + increment)
    
    def generate(self, json_data, output_path=None):
        """生成Excel报告"""
        if self.config.get("create_from_blank"):
            return self._generate_from_blank(json_data, output_path)

        template_path = self.config.get("template_file", "template/report_template.xlsx")

        if not os.path.isabs(template_path) and not os.path.exists(template_path):
            possible_paths = [
                template_path,
                os.path.join(self.database_path, "templates", template_path),
                os.path.join(os.path.dirname(__file__), template_path),
                os.path.join("template", template_path),
                os.path.join("arch", template_path),
                os.path.join(self.database_path, template_path),
            ]
            for path in possible_paths:
                if os.path.exists(path):
                    template_path = path
                    break

        if not os.path.exists(template_path):
            raise FileNotFoundError(f"模板文件不存在: {template_path}")

        wb = load_workbook(template_path)
        sheet_name = self.config.get("sheet_name")
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        self._clear_template_sheet(ws)

        for data_cell in self.config.get("data_cells", []):
            coord = data_cell.get("range") or data_cell.get("cell")
            if not coord:
                continue
            target_coord = self._apply_cell_range(ws, coord)
            cell = ws[target_coord]
            value = self._resolve_data_value(data_cell, json_data)
            cell.value = value
            style_name = data_cell.get("style")
            if style_name:
                self._apply_style(cell, style_name)

        self._apply_page_setup(ws, self.config.get("page_setup", {}))

        suffix = (
            self.template_id if self.template_id and self.template_id != "default_template" else None
        )
        output_path = self._build_output_path(json_data, output_path, suffix=suffix)
        wb.save(output_path)
        return output_path

    def _generate_from_blank(self, json_data, output_path=None):
        wb = Workbook()
        ws = wb.active
        ws.title = self.config.get("sheet_name", "Sheet1")

        for col, width in self.config.get("column_widths", {}).items():
            ws.column_dimensions[col].width = width

        for row, height in self.config.get("row_heights", {}).items():
            try:
                ws.row_dimensions[int(row)].height = height
            except (TypeError, ValueError):
                continue

        for static_cell in self.config.get("static_cells", []):
            coord = self._apply_cell_range(ws, static_cell.get("range"))
            cell = ws[coord]
            style_name = static_cell.get("style")
            if style_name:
                self._apply_style(cell, style_name)
            if "value" in static_cell:
                cell.value = static_cell.get("value", "")

        for data_cell in self.config.get("data_cells", []):
            coord = self._apply_cell_range(ws, data_cell.get("range"))
            cell = ws[coord]
            value = self._resolve_data_value(data_cell, json_data)
            cell.value = value
            style_name = data_cell.get("style")
            if style_name:
                self._apply_style(cell, style_name)

        self._apply_page_setup(ws, self.config.get("page_setup", {}))

        suffix = (
            self.template_id if self.template_id and self.template_id != "default_template" else None
        )
        output_path = self._build_output_path(
            json_data,
            output_path,
            suffix=suffix,
            use_date_folder=True,
        )
        wb.save(output_path)
        return output_path

    def _build_output_path(self, json_data, output_path, *, suffix=None, use_date_folder=False):
        if output_path:
            return output_path

        output_dir = self.config.get("output_directory", "excel")
        if not os.path.isabs(output_dir):
            output_dir = os.path.join(self.database_path, output_dir)

        if use_date_folder:
            date_str = self._get_data_value(json_data, "基本信息.检查时间") or datetime.now().strftime(
                "%Y-%m-%d"
            )
            date_str = date_str.replace("/", "-").replace("\\", "-")
            output_dir = os.path.join(output_dir, self._sanitize_path_component(date_str))

        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        patient_name = self._get_data_value(json_data, "基本信息.姓名") or "unknown"
        patient_id = self._get_data_value(json_data, "基本信息.ID") or "unknown"

        filename_parts = [
            self._sanitize_path_component(patient_name),
            self._sanitize_path_component(patient_id),
            timestamp,
        ]
        if suffix:
            filename_parts.append(self._sanitize_path_component(str(suffix)))

        filename = "_".join(filename_parts) + ".xlsx"
        return os.path.join(output_dir, filename)

    def _apply_cell_range(self, worksheet, cell_range):
        if not cell_range:
            raise ValueError("单元格范围不能为空")
        if ":" in cell_range:
            worksheet.merge_cells(cell_range)
            start = cell_range.split(":")[0]
        else:
            start = cell_range
        return start

    def _resolve_data_value(self, cell_config, data):
        paths = cell_config.get("paths")
        value = ""

        if paths:
            for path in paths:
                value = self._get_data_value(data, path)
                if not self._is_empty(value):
                    break
        else:
            value = self._get_data_value(data, cell_config.get("data_path", ""))

        value_type = cell_config.get("value_type")
        if value_type == "main_with_note":
            secondary = self._get_data_value(data, cell_config.get("secondary_path", ""))
            value = self._merge_main_and_secondary(value, secondary)

        format_type = cell_config.get("format")
        if isinstance(value, dict):
            if format_type == "list":
                if "检查结论" in value and isinstance(value["检查结论"], list):
                    value = value["检查结论"]
                else:
                    value = [str(v) for v in value.values() if not self._is_empty(v)]
            elif "检查所见" in value and isinstance(value["检查所见"], str):
                value = value["检查所见"]
            else:
                value = ""
        if format_type:
            value = self._format_value(value, format_type)

        if self._is_empty(value):
            value = cell_config.get(
                "default",
                self.config.get("empty_value_handling", {}).get("default", ""),
            )

        return value

    def _merge_main_and_secondary(self, main, secondary):
        if self._is_empty(main) and self._is_empty(secondary):
            return ""
        if self._is_empty(secondary):
            return main
        if self._is_empty(main):
            return secondary
        return f"{main}（{secondary}）"

    def _get_merged_master_cell(self, worksheet, coord):
        for merged_range in worksheet.merged_cells.ranges:
            if coord in merged_range:
                start = str(merged_range).split(":")[0]
                return worksheet[start]
        return worksheet[coord]

    def _clear_template_sheet(self, worksheet):
        pass

    def _sanitize_path_component(self, value: str) -> str:
        if not value:
            return "unknown"
        return re.sub(r'[\\\\/:*?"<>|]', "_", value)
    
    def _process_section(self, ws, worksheet_name, section, data):
        """处理一个section"""
        section_type = section.get("type", "fixed")
        
        # 检查条件
        condition = section.get("condition", {})
        if condition.get("skip_if_empty", False):
            check_key = condition.get("check_empty")
            if check_key and self._is_section_empty(data, check_key):
                return
        
        # 确定起始行
        if section_type == "fixed":
            start_row = section.get("start_row", 1)
            self._set_current_row(worksheet_name, start_row)
        elif section_type == "conditional" or section_type == "dynamic":
            dynamic_config = section.get("dynamic_start_row", {})
            base = dynamic_config.get("base", 1)
            previous_section = dynamic_config.get("previous_section")
            if previous_section:
                # 从上一个section的结束行开始
                start_row = self._get_current_row(worksheet_name)
            else:
                start_row = base
            self._set_current_row(worksheet_name, start_row)
        
        # 处理layout
        layout = section.get("layout", [])
        if layout:
            for layout_item in layout:
                current_row = self._get_current_row(worksheet_name)
                row_offset = layout_item.get("row_offset", 0)
                target_row = current_row + row_offset
                
                self._process_layout_item(ws, worksheet_name, layout_item, data, target_row)
            
            # 更新当前行到最后处理的行的最大行号
            max_row_used = max([self._get_current_row(worksheet_name) + item.get("row_offset", 0) 
                               for item in layout], default=self._get_current_row(worksheet_name))
            max_row_used += layout[-1].get("spacer", {}).get("height", 0) if layout else 0
            self._set_current_row(worksheet_name, max_row_used)
        else:
            # 简单的cells配置（向后兼容）
            cells = section.get("cells", [])
            for cell_config in cells:
                self._process_cell(ws, cell_config, data)
    
    def _process_layout_item(self, ws, worksheet_name, layout_item, data, row):
        """处理布局项"""
        item_type = layout_item.get("type")
        
        if item_type == "section_title":
            merge_range = self._resolve_cell_range(layout_item.get("merge_range", ""), row)
            text = layout_item.get("text", "")
            style_name = layout_item.get("style", "section_title")
            
            ws.merge_cells(merge_range)
            cell = ws[merge_range.split(':')[0]]
            cell.value = text
            self._apply_style(cell, style_name)
        
        elif item_type == "header_row" or item_type == "data_row":
            cells_config = layout_item.get("cells", [])
            for cell_config in cells_config:
                merge_range = self._resolve_cell_range(cell_config.get("merge_range", ""), row)
                data_path = cell_config.get("data_path")
                text = cell_config.get("text")
                style_name = cell_config.get("style", "data_cell")
                
                if merge_range:
                    ws.merge_cells(merge_range)
                    cell = ws[merge_range.split(':')[0]]
                else:
                    cell = ws[cell_config.get("cell", "A1")]
                
                # 设置值
                if data_path:
                    value = self._get_data_value(data, data_path)
                    format_type = cell_config.get("format", "text")
                    value = self._format_value(value, format_type)
                    if self._is_empty(value):
                        value = self.config.get("empty_value_handling", {}).get("default", "")
                    cell.value = value
                elif text:
                    cell.value = text
                
                self._apply_style(cell, style_name)
        
        elif item_type == "spacer":
            height = layout_item.get("height", 1)
            self._increment_row(worksheet_name, height)
    
    def _process_cell(self, ws, cell_config, data):
        """处理单个单元格（向后兼容）"""
        cell_ref = cell_config.get("cell")
        data_path = cell_config.get("data_path")
        format_type = cell_config.get("format", "text")
        
        if cell_ref and data_path:
            value = self._get_data_value(data, data_path)
            value = self._format_value(value, format_type)
            if self._is_empty(value):
                value = self.config.get("empty_value_handling", {}).get("default", "")
            cell = ws[cell_ref]
            if isinstance(cell, MergedCell):
                cell = self._get_merged_master_cell(ws, cell_ref)
            cell.value = value
            style_name = cell_config.get("style")
            if style_name:
                self._apply_style(cell, style_name)
    
    def _apply_page_setup(self, ws, page_setup_config):
        """应用页面设置"""
        if not page_setup_config:
            return
        
        # 方向
        orientation = page_setup_config.get("orientation", "portrait")
        ws.page_setup.orientation = orientation
        
        # 适应宽度
        if page_setup_config.get("fitToWidth"):
            ws.page_setup.fitToWidth = page_setup_config.get("fitToWidth")
        
        # 适应高度
        fit_to_height = page_setup_config.get("fitToHeight")
        if fit_to_height is not None:
            ws.page_setup.fitToHeight = fit_to_height
        
        # 页边距
        margins = page_setup_config.get("margins", {})
        if margins:
            ws.page_margins.left = margins.get("left", 0.1)
            ws.page_margins.right = margins.get("right", 0.1)
            ws.page_margins.top = margins.get("top", 0.1)
            ws.page_margins.bottom = margins.get("bottom", 0.1)
            ws.page_margins.header = margins.get("header", 0.1)
            ws.page_margins.footer = margins.get("footer", 0.1)
        
        # 打印区域
        print_area_config = page_setup_config.get("print_area", {})
        if print_area_config and print_area_config.get("dynamic"):
            current_row = self._get_current_row(ws.title)
            base_col = print_area_config.get("base_col", "A")
            end_col = print_area_config.get("end_col", "M")
            ws.print_area = f"{base_col}1:{end_col}{current_row}"


if __name__ == "__main__":
    # 测试代码
    generator = ExcelGenerator(database_path="vest_database", template_id="default_template")
    test_data = {
        "基本信息": {
            "ID": "12345",
            "姓名": "测试患者",
            "性别": "男",
            "出生日期": "1990/01/01",
            "检查时间": "2025/01/01",
            "检查设备": "测试设备"
        },
        "自发性眼震": {
            "自发性眼震模式": "右跳性眼震",
            "自发性眼震速度": "5.2",
            "自发性眼震固视抑制": "正常",
            "自发性眼震检查结果": "异常"
        }
    }
    
    output = generator.generate(test_data)
    print(f"Excel文件已生成: {output}")

